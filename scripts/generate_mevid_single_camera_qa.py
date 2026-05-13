from __future__ import annotations

import argparse
import json
import random
import re
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_XLSX = ROOT / "agent" / "test" / "data" / "agent_test_mevid.xlsx"
DEFAULT_RAW_DIR = ROOT / "tests" / "mevid_xlsx"
DEFAULT_OUT_XLSX = ROOT / "agent" / "test" / "data" / "mevid_single_camera_qa_cleaned.xlsx"
DEFAULT_REPORT_JSON = ROOT / "agent" / "test" / "data" / "mevid_single_camera_qa_cleaned_report.json"

HEADERS = [
    "video_id",
    "question",
    "recall_challenge",
    "expected_answer",
    "expected_time",
    "difficulty",
    "video_quality",
]

CAMERA_RE = re.compile(r"\bG\d{3}\b")
TIME_RANGE_RE = re.compile(r"^\s*(\d+:\d{1,2})\s*-\s*(\d+:\d{1,2})\s*$")


@dataclass(frozen=True)
class BaseCase:
    video_id: str
    camera_id: str
    slot: str
    question: str
    recall_challenge: str
    expected_time: str
    difficulty: str
    video_quality: str
    category: str
    direction: str
    appearance_phrase: str


def _slot_from_video_id(video_id: str) -> str:
    parts = video_id.split(".")
    if len(parts) < 2:
        return "unknown"
    m = re.match(r"(\d{2})-(\d{2})", parts[1])
    return f"{m.group(1)}-{m.group(2)}" if m else parts[1]


def _camera_from_text(*values: str) -> str:
    for value in values:
        match = CAMERA_RE.search(value or "")
        if match:
            return match.group(0)
    return ""


def _valid_time_range(value: Any) -> str:
    text = str(value or "").strip()
    return text if TIME_RANGE_RE.match(text) else ""


def _category(question: str, expected_answer: str) -> str:
    q = question.lower()
    if expected_answer.strip().lower() == "no":
        return "negative"
    if "exit" in q or "leaving" in q or "leave" in q:
        return "event"
    if "appear in camera" in q and "then appear" in q:
        return "cross_camera"
    if "visible" in q or "wearing" in q or "with" in q:
        return "appearance"
    return "unknown"


def _appearance_phrase(question: str, recall_challenge: str) -> str:
    q = question.strip()
    patterns = [
        r"person with (.+?) visible",
        r"person wearing (.+?) in camera",
        r"person with (.+?) in camera",
    ]
    for pattern in patterns:
        match = re.search(pattern, q, flags=re.IGNORECASE)
        if match:
            return _clean_phrase(match.group(1))
    challenge = _clean_phrase((recall_challenge or "").split(";")[0].split(".")[0])
    return challenge or "the described appearance"


def _direction(question: str, recall_challenge: str) -> str:
    text = f"{question} {recall_challenge}".lower()
    candidates = [
        "top left",
        "top right",
        "bottom left",
        "bottom right",
        "left",
        "right",
        "up",
        "down",
        "top",
        "bottom",
    ]
    for direction in candidates:
        if re.search(rf"\b{re.escape(direction)}\b", text):
            if direction == "top":
                return "up"
            if direction == "bottom":
                return "down"
            return direction
    return ""


def _clean_phrase(text: str) -> str:
    text = (text or "").strip()
    text = re.sub(r"\s+", " ", text)
    text = text.strip(" .,;:-")
    return text


def _lower_first(text: str) -> str:
    text = text.strip()
    if not text:
        return text
    return text[0].lower() + text[1:]


def _question_prefix(case: BaseCase) -> str:
    return f"In the {case.slot} video for camera {case.camera_id}, around {case.expected_time}, "


def _rewrite_positive(case: BaseCase) -> list[Any]:
    prefix = _question_prefix(case)
    if case.category == "event" and case.direction:
        question = f"{prefix}did a person exit from the {case.direction} side?"
    elif case.category == "appearance":
        question = f"{prefix}is there a person with {case.appearance_phrase} visible?"
    else:
        question = f"{prefix}{_lower_first(case.question).rstrip('?')}?"
    return [
        case.video_id,
        question,
        case.recall_challenge,
        "yes",
        case.expected_time,
        case.difficulty,
        case.video_quality,
    ]


def _wrong_direction(direction: str) -> str:
    mapping = {
        "left": "right",
        "right": "left",
        "up": "down",
        "down": "up",
        "top left": "right",
        "top right": "left",
        "bottom left": "right",
        "bottom right": "left",
    }
    return mapping.get(direction, "left")


def _short_desc(phrase: str) -> str:
    phrase = _clean_phrase(phrase)
    parts = [p.strip() for p in re.split(r",|;|\band\b", phrase) if p.strip()]
    return parts[0] if parts else phrase


def _negative_appearance_for(case: BaseCase, candidates: list[BaseCase]) -> str:
    true_desc = _short_desc(case.appearance_phrase).lower()
    for candidate in candidates:
        if candidate.category != "appearance":
            continue
        if candidate.video_id != case.video_id or candidate.camera_id != case.camera_id:
            continue
        other = _short_desc(candidate.appearance_phrase)
        if other and other.lower() != true_desc:
            return other
    fallbacks = [
        "bright red jacket",
        "white hoodie",
        "light blue top",
        "yellow coat",
        "blue and black plaid shirt",
        "black backpack",
    ]
    for fallback in fallbacks:
        if fallback.lower() not in case.recall_challenge.lower() and fallback.lower() not in true_desc:
            return fallback
    return "a clearly different outfit"


def _make_negative_appearance(case: BaseCase, candidates: list[BaseCase]) -> list[Any]:
    false_desc = _negative_appearance_for(case, candidates)
    question = f"{_question_prefix(case)}is there a person with {false_desc} visible?"
    recall = (
        "Negative single-camera appearance check. "
        f"Target window contains: {case.recall_challenge}. "
        f"The question asks for different appearance: {false_desc}."
    )
    return [
        case.video_id,
        question,
        recall,
        "no",
        case.expected_time,
        "medium",
        case.video_quality,
    ]


def _make_negative_event(case: BaseCase) -> list[Any]:
    wrong = _wrong_direction(case.direction)
    question = f"{_question_prefix(case)}did a person exit from the {wrong} side?"
    recall = (
        "Negative single-camera event check. "
        f"Target window contains a person exiting from the {case.direction} side, not the {wrong} side. "
        f"Appearance/context: {case.recall_challenge}."
    )
    return [
        case.video_id,
        question,
        recall,
        "no",
        case.expected_time,
        "medium",
        case.video_quality,
    ]


def _load_base_cases(source_xlsx: Path) -> list[BaseCase]:
    wb = load_workbook(source_xlsx, read_only=True, data_only=True)
    ws = wb["Part1"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h or "").strip() for h in rows[0]]
    idx = {name: i for i, name in enumerate(headers)}
    cases: list[BaseCase] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    for row in rows[1:]:
        video_id = str(row[idx["video_id"]] or "").strip()
        question = str(row[idx["question"]] or "").strip()
        expected_answer = str(row[idx["expected_answer"]] or "").strip().lower()
        expected_time = _valid_time_range(row[idx["expected_time"]])
        if not video_id or expected_answer != "yes" or not expected_time:
            continue
        camera_id = _camera_from_text(video_id, question)
        if not camera_id:
            continue
        category = _category(question, expected_answer)
        if category not in {"appearance", "event"}:
            continue
        recall_challenge = str(row[idx["recall_challenge"]] or "").strip()
        direction = _direction(question, recall_challenge)
        if category == "event" and not direction:
            continue
        appearance = _appearance_phrase(question, recall_challenge)
        dedupe_key = (
            video_id,
            camera_id,
            expected_time,
            category,
            appearance.lower() if category == "appearance" else direction.lower(),
        )
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        cases.append(
            BaseCase(
                video_id=video_id,
                camera_id=camera_id,
                slot=_slot_from_video_id(video_id),
                question=question,
                recall_challenge=recall_challenge,
                expected_time=expected_time,
                difficulty=str(row[idx["difficulty"]] or "medium").strip() or "medium",
                video_quality=str(row[idx["video_quality"]] or "high").strip() or "high",
                category=category,
                direction=direction,
                appearance_phrase=appearance,
            )
        )
    return cases


def _raw_manifest(raw_dir: Path) -> dict[str, Any]:
    manifest: dict[str, Any] = {"raw_dir": str(raw_dir), "files": []}
    for path in sorted(raw_dir.glob("*.xlsx")):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb["Sightings"] if "Sightings" in wb.sheetnames else wb.worksheets[0]
            count = 0
            cameras: Counter[str] = Counter()
            for row in ws.iter_rows(min_row=4, values_only=True):
                camera = _camera_from_text(str(row[2] or "") if len(row) > 2 else "")
                appearance = str(row[6] or "").strip() if len(row) > 6 else ""
                if camera and appearance:
                    count += 1
                    cameras[camera] += 1
            manifest["files"].append(
                {
                    "name": path.name,
                    "sighting_rows": count,
                    "cameras": dict(sorted(cameras.items())),
                }
            )
        except Exception as exc:  # pragma: no cover - report only
            manifest["files"].append({"name": path.name, "error": str(exc)})
    return manifest


def _select_balanced(
    cases: list[BaseCase],
    *,
    max_per_video_camera: int,
    include_negative: bool,
    seed: int,
) -> list[list[Any]]:
    rng = random.Random(seed)
    by_video_camera: dict[tuple[str, str], list[BaseCase]] = defaultdict(list)
    for case in cases:
        by_video_camera[(case.video_id, case.camera_id)].append(case)

    output: list[list[Any]] = []
    all_cases = list(cases)
    for key in sorted(by_video_camera):
        group = by_video_camera[key]
        appearances = [c for c in group if c.category == "appearance"]
        events = [c for c in group if c.category == "event"]
        rng.shuffle(appearances)
        rng.shuffle(events)
        pos = appearances[: max_per_video_camera // 2] + events[: max_per_video_camera // 2]
        if not pos:
            pos = group[:max_per_video_camera]
        for case in pos[:max_per_video_camera]:
            output.append(_rewrite_positive(case))
            if include_negative:
                if case.category == "appearance":
                    output.append(_make_negative_appearance(case, all_cases))
                elif case.category == "event":
                    output.append(_make_negative_event(case))
    return output


def _write_workbook(rows: list[list[Any]], out_xlsx: Path) -> None:
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Part1"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    legend = wb.create_sheet("Legend")
    legend.append(["kind", "template"])
    legend.append(["appearance_positive", "In the [slot] video for camera [Gxxx], around [mm:ss-mm:ss], is there a person with [desc] visible?"])
    legend.append(["appearance_negative", "Same window, ask for a different appearance; expected no."])
    legend.append(["event_positive", "In the [slot] video for camera [Gxxx], around [mm:ss-mm:ss], did a person exit from the [side] side?"])
    legend.append(["event_negative", "Same window, ask for the wrong side; expected no."])
    wb.save(out_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate cleaned single-camera MEVID QA cases.")
    parser.add_argument("--source-xlsx", type=Path, default=DEFAULT_SOURCE_XLSX)
    parser.add_argument("--raw-dir", type=Path, default=DEFAULT_RAW_DIR)
    parser.add_argument("--out-xlsx", type=Path, default=DEFAULT_OUT_XLSX)
    parser.add_argument("--report-json", type=Path, default=DEFAULT_REPORT_JSON)
    parser.add_argument("--max-per-video-camera", type=int, default=8)
    parser.add_argument("--seed", type=int, default=13)
    parser.add_argument("--no-negative", action="store_true")
    args = parser.parse_args()

    base_cases = _load_base_cases(args.source_xlsx)
    rows = _select_balanced(
        base_cases,
        max_per_video_camera=max(1, args.max_per_video_camera),
        include_negative=not args.no_negative,
        seed=args.seed,
    )
    _write_workbook(rows, args.out_xlsx)

    counts = Counter()
    by_video_camera = Counter()
    for row in rows:
        expected = str(row[3]).lower()
        q = str(row[1]).lower()
        if expected == "no":
            counts["negative"] += 1
        elif "exit" in q:
            counts["event"] += 1
        else:
            counts["appearance"] += 1
        by_video_camera[(str(row[0]), _camera_from_text(str(row[0]), str(row[1])))] += 1

    report = {
        "source_xlsx": str(args.source_xlsx),
        "out_xlsx": str(args.out_xlsx),
        "base_positive_cases": len(base_cases),
        "generated_cases": len(rows),
        "category_counts": dict(counts),
        "positive_count": sum(1 for row in rows if str(row[3]).lower() == "yes"),
        "negative_count": sum(1 for row in rows if str(row[3]).lower() == "no"),
        "video_camera_counts": {
            f"{video_id}|{camera_id}": count
            for (video_id, camera_id), count in sorted(by_video_camera.items())
        },
        "raw_manifest": _raw_manifest(args.raw_dir),
    }
    args.report_json.parent.mkdir(parents=True, exist_ok=True)
    args.report_json.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Wrote {len(rows)} cases -> {args.out_xlsx}")
    print(f"Report -> {args.report_json}")
    print(json.dumps({k: report[k] for k in ["base_positive_cases", "generated_cases", "category_counts"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
