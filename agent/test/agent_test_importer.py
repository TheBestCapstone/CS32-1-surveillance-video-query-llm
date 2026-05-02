from __future__ import annotations

import argparse
import json
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_TEST_DATA_ROOT = ROOT_DIR / "agent" / "test" / "data"
DEFAULT_XLSX_PATH = DEFAULT_TEST_DATA_ROOT / "agent_test.xlsx"
# Default output: agent/test/data/imported/
DEFAULT_IMPORT_OUTPUT_DIR = DEFAULT_TEST_DATA_ROOT / "imported"
DEFAULT_IMPORT_INCLUDE_SHEETS = ("Part4",)
# Legacy path kept for callers that referenced the older default.
DEFAULT_OUTPUT_DIR = ROOT_DIR / "agent" / "test" / "generated"
DEFAULT_DB_PATH = DEFAULT_OUTPUT_DIR / "agent_test_eval.sqlite"
SQLITE_EVAL_FILENAME = DEFAULT_DB_PATH.name
DEFAULT_IMPORT_SQLITE_PATH = DEFAULT_IMPORT_OUTPUT_DIR / SQLITE_EVAL_FILENAME
DEFAULT_JSON_PATH = DEFAULT_OUTPUT_DIR / "agent_test_normalized.json"
DEFAULT_REPORT_PATH = DEFAULT_OUTPUT_DIR / "agent_test_import_report.json"
DEFAULT_RETRIEVAL_VIEW = DEFAULT_OUTPUT_DIR / "agent_test_retrieval_eval.json"
DEFAULT_E2E_VIEW = DEFAULT_OUTPUT_DIR / "agent_test_e2e_eval.json"
DEFAULT_GENERATION_VIEW = DEFAULT_OUTPUT_DIR / "agent_test_generation_eval.json"
SCHEMA_VERSION = "v1_agent_test_eval_import"
TABLE_NAME = "agent_eval_cases"


CREATE_TABLE_SQL = f"""
CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
    case_id TEXT PRIMARY KEY,
    source_sheet TEXT NOT NULL,
    source_row INTEGER,
    source_format TEXT NOT NULL,
    schema_version TEXT NOT NULL,
    video_id TEXT,
    question TEXT NOT NULL,
    question_language TEXT,
    recall_challenge TEXT,
    expected_answer_raw TEXT,
    expected_answer_label TEXT,
    expected_time_raw TEXT,
    expected_start_sec REAL,
    expected_end_sec REAL,
    expected_time_is_approx INTEGER,
    difficulty_raw TEXT,
    difficulty_level TEXT,
    video_quality_raw TEXT,
    video_quality_level TEXT,
    reference_answer TEXT,
    retrieval_ready INTEGER,
    e2e_ready INTEGER,
    generation_ready INTEGER,
    metadata_json TEXT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP
);
"""


@dataclass
class AgentTestImportConfig:
    xlsx_path: Path = DEFAULT_XLSX_PATH
    output_dir: Path = DEFAULT_IMPORT_OUTPUT_DIR
    sqlite_path: Path = DEFAULT_IMPORT_SQLITE_PATH
    normalized_json_path: Path = DEFAULT_JSON_PATH
    report_json_path: Path = DEFAULT_REPORT_PATH
    retrieval_view_path: Path = DEFAULT_RETRIEVAL_VIEW
    e2e_view_path: Path = DEFAULT_E2E_VIEW
    generation_view_path: Path = DEFAULT_GENERATION_VIEW
    reset_existing: bool = False
    include_sheets: list[str] | None = None


class AgentTestImportError(RuntimeError):
    pass


class AgentTestDatasetImporter:
    def __init__(self, config: AgentTestImportConfig) -> None:
        self.config = config

    def build(self) -> dict[str, Any]:
        if not self.config.xlsx_path.exists():
            raise AgentTestImportError(f"Excel file not found: {self.config.xlsx_path}")

        self.config.output_dir.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(self.config.xlsx_path, data_only=True)
        cases = self._parse_workbook(workbook)
        self._write_json(self.config.normalized_json_path, cases)
        self._write_views(cases)
        self._write_sqlite(cases)
        report = self._build_report(cases)
        self._write_json(self.config.report_json_path, report)
        return report

    def _parse_workbook(self, workbook) -> list[dict[str, Any]]:
        cases: list[dict[str, Any]] = []
        include_sheets = {str(item).strip() for item in (self.config.include_sheets or []) if str(item).strip()}
        for sheet in workbook.worksheets:
            title = str(sheet.title)
            if include_sheets and title not in include_sheets:
                continue
            if title in {"Part2", "Part6"}:
                continue
            if title == "Part3":
                cases.extend(self._parse_part3(sheet))
            else:
                cases.extend(self._parse_standard_sheet(sheet))
        return cases

    def _parse_standard_sheet(self, sheet) -> list[dict[str, Any]]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        current_video_id: str | None = None
        out: list[dict[str, Any]] = []
        for row_idx, row in enumerate(rows[1:], start=2):
            values = [self._clean_cell(item) for item in row]
            if not any(values):
                continue
            video_id = values[0] or current_video_id
            if values[0]:
                current_video_id = values[0]
            question = values[1]
            if not question:
                continue
            case = self._make_case(
                source_sheet=sheet.title,
                source_row=row_idx,
                video_id=video_id,
                question=question,
                recall_challenge=values[2] if len(values) > 2 else None,
                expected_answer=values[3] if len(values) > 3 else None,
                expected_time=values[4] if len(values) > 4 else None,
                difficulty=values[5] if len(values) > 5 else None,
                video_quality=values[6] if len(values) > 6 else None,
                metadata={"raw_row": values},
            )
            out.append(case)
        return out

    def _parse_part3(self, sheet) -> list[dict[str, Any]]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        out: list[dict[str, Any]] = []
        current_video_id: str | None = None
        current_difficulty: str | None = None
        pending_question: str | None = None
        pending_notes: list[str] = []

        def flush_pending(row_idx: int) -> None:
            nonlocal pending_question, pending_notes
            if not pending_question:
                pending_notes = []
                return
            challenge = " ".join(note for note in pending_notes if note).strip() or None
            out.append(
                self._make_case(
                    source_sheet=sheet.title,
                    source_row=row_idx,
                    video_id=current_video_id,
                    question=pending_question,
                    recall_challenge=challenge,
                    expected_answer=None,
                    expected_time=None,
                    difficulty=current_difficulty,
                    video_quality=None,
                    metadata={"raw_notes": list(pending_notes), "parse_mode": "part3_block"},
                )
            )
            pending_question = None
            pending_notes = []

        for row_idx, row in enumerate(rows[1:], start=2):
            values = [self._clean_cell(item) for item in row]
            if not any(values):
                continue

            if values[0] and values[2]:
                flush_pending(row_idx - 1)
                out.append(
                    self._make_case(
                        source_sheet=sheet.title,
                        source_row=row_idx,
                        video_id=values[0],
                        question=values[1],
                        recall_challenge=None,
                        expected_answer=values[2],
                        expected_time=values[3],
                        difficulty=values[4],
                        video_quality=values[5],
                        metadata={"raw_row": values, "parse_mode": "part3_structured"},
                    )
                )
                continue

            if values[0]:
                current_video_id = values[0]
            cell = values[1]
            if not cell:
                continue
            if self._looks_like_difficulty(cell):
                flush_pending(row_idx - 1)
                current_difficulty = cell
                continue
            if cell.startswith("问题"):
                flush_pending(row_idx - 1)
                pending_question = self._strip_prefix(cell, "问题")
                continue
            pending_notes.append(cell)

        flush_pending(sheet.max_row)
        return out

    def _make_case(
        self,
        *,
        source_sheet: str,
        source_row: int,
        video_id: str | None,
        question: str | None,
        recall_challenge: str | None,
        expected_answer: str | None,
        expected_time: str | None,
        difficulty: str | None,
        video_quality: str | None,
        metadata: dict[str, Any],
    ) -> dict[str, Any]:
        question_text = self._normalize_question(question)
        if not question_text:
            raise AgentTestImportError(f"Empty question at {source_sheet}:{source_row}")
        answer_label = self._normalize_answer_label(expected_answer)
        time_info = self._parse_time_range(expected_time)
        difficulty_level = self._normalize_difficulty(difficulty)
        video_quality_level = self._normalize_video_quality(video_quality)
        question_language = self._infer_language(question_text)
        reference_answer = self._build_reference_answer(
            video_id=video_id,
            answer_label=answer_label,
            expected_time_raw=expected_time,
            question_language=question_language,
        )
        reference_scene_description = self._build_reference_scene_description(
            recall_challenge=recall_challenge,
            question=question_text,
            question_language=question_language,
        )
        reference_answer_rich = self._build_reference_answer_rich(
            video_id=video_id,
            answer_label=answer_label,
            expected_time_raw=expected_time,
            scene_description=reference_scene_description,
            question_language=question_language,
        )
        retrieval_ready = 1 if question_text and video_id else 0
        e2e_ready = 1 if retrieval_ready and answer_label in {"yes", "no"} else 0
        generation_ready = 1 if e2e_ready and reference_answer else 0
        return {
            "case_id": f"{source_sheet.upper()}_{source_row:04d}",
            "source_sheet": source_sheet,
            "source_row": source_row,
            "source_format": "agent_test_xlsx",
            "schema_version": SCHEMA_VERSION,
            "video_id": video_id,
            "question": question_text,
            "question_language": question_language,
            "recall_challenge": recall_challenge,
            "expected_answer_raw": expected_answer,
            "expected_answer_label": answer_label,
            "expected_time_raw": expected_time,
            "expected_start_sec": time_info["start_sec"],
            "expected_end_sec": time_info["end_sec"],
            "expected_time_is_approx": 1 if time_info["is_approx"] else 0,
            "difficulty_raw": difficulty,
            "difficulty_level": difficulty_level,
            "video_quality_raw": video_quality,
            "video_quality_level": video_quality_level,
            "reference_answer": reference_answer,
            "reference_scene_description": reference_scene_description,
            "reference_answer_rich": reference_answer_rich,
            "retrieval_ready": retrieval_ready,
            "e2e_ready": e2e_ready,
            "generation_ready": generation_ready,
            "metadata_json": metadata,
        }

    @staticmethod
    def _clean_cell(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _strip_prefix(text: str, prefix: str) -> str:
        cleaned = text.strip()
        if cleaned.startswith(prefix + "："):
            return cleaned[len(prefix) + 1 :].strip()
        if cleaned.startswith(prefix + ":"):
            return cleaned[len(prefix) + 1 :].strip()
        if cleaned.startswith(prefix):
            return cleaned[len(prefix) :].strip()
        return cleaned

    @staticmethod
    def _normalize_question(text: str | None) -> str:
        cleaned = (text or "").strip()
        cleaned = AgentTestDatasetImporter._strip_prefix(cleaned, "问题")
        return cleaned.strip("“”\" ").strip()

    @staticmethod
    def _infer_language(text: str) -> str:
        return "zh" if re.search(r"[\u4e00-\u9fff]", text or "") else "en"

    @staticmethod
    def _looks_like_difficulty(text: str) -> bool:
        value = (text or "").strip().lower()
        return value in {
            "1",
            "2",
            "3",
            "1.0",
            "2.0",
            "3.0",
            "easy",
            "medium",
            "hard",
            "简单",
            "中等",
            "困难",
        }

    @staticmethod
    def _normalize_answer_label(value: str | None) -> str:
        text = (value or "").strip().lower()
        if text in {"yes", "y", "有", "是"}:
            return "yes"
        if text in {"no", "n", "没有", "无", "否"}:
            return "no"
        return "unknown"

    @staticmethod
    def _normalize_difficulty(value: str | None) -> str:
        text = (value or "").strip().lower()
        mapping = {
            "1": "easy",
            "1.0": "easy",
            "easy": "easy",
            "简单": "easy",
            "2": "medium",
            "2.0": "medium",
            "medium": "medium",
            "中等": "medium",
            "3": "hard",
            "3.0": "hard",
            "hard": "hard",
            "困难": "hard",
        }
        return mapping.get(text, "unknown")

    @staticmethod
    def _normalize_video_quality(value: str | None) -> str:
        text = (value or "").strip().lower()
        mapping = {
            "high": "high",
            "较高": "high",
            "medium": "medium",
            "medium-high": "medium",
            "medium low": "medium",
            "medium_low": "medium",
            "中等": "medium",
            "中等偏低": "medium",
            "low": "low",
            "较低": "low",
        }
        return mapping.get(text, "unknown")

    @staticmethod
    def _parse_time_range(value: str | None) -> dict[str, Any]:
        raw = (value or "").strip()
        if not raw or raw.lower() in {"n/a", "na", "none", "null"}:
            return {"start_sec": None, "end_sec": None, "is_approx": False}

        normalized = (
            raw.replace("–", "-")
            .replace("—", "-")
            .replace("－", "-")
            .replace("〜", "-")
            .replace("~", "-")
        )
        is_approx = "附近" in normalized or "around" in normalized.lower()
        normalized = normalized.replace("附近", "").replace("around", "").strip()

        parts = [part.strip() for part in normalized.split("-") if part.strip()]
        if len(parts) >= 2:
            start_sec = AgentTestDatasetImporter._parse_single_time(parts[0])
            end_sec = AgentTestDatasetImporter._parse_single_time(parts[1])
            return {"start_sec": start_sec, "end_sec": end_sec, "is_approx": is_approx}

        point = AgentTestDatasetImporter._parse_single_time(normalized)
        return {"start_sec": point, "end_sec": point, "is_approx": True}

    @staticmethod
    def _parse_single_time(value: str | None) -> float | None:
        text = (value or "").strip()
        if not text:
            return None
        if ":" in text:
            parts = text.split(":")
            try:
                nums = [float(part) for part in parts]
            except Exception:
                return None
            if len(nums) == 3:
                return nums[0] * 3600 + nums[1] * 60 + nums[2]
            if len(nums) == 2:
                return nums[0] * 60 + nums[1]
            return None
        try:
            return float(text)
        except Exception:
            return None

    @staticmethod
    def _build_reference_answer(
        *,
        video_id: str | None,
        answer_label: str,
        expected_time_raw: str | None,
        question_language: str,
    ) -> str | None:
        if answer_label == "yes":
            if question_language == "zh":
                if expected_time_raw:
                    return f"有，对应视频为 {video_id}，参考时间为 {expected_time_raw}。"
                return f"有，对应视频为 {video_id}。"
            if expected_time_raw:
                return f"Yes. The relevant clip is in {video_id}, around {expected_time_raw}."
            return f"Yes. The relevant clip is in {video_id}."
        if answer_label == "no":
            return "没有匹配片段。" if question_language == "zh" else "No matching clip is expected."
        return None

    # ----- challenge.md: rich reference for RAGAS recall/precision/faithfulness -----

    _QUESTION_EN_STOP_PREFIXES = (
        "is there a clip of ",
        "is there a clip showing ",
        "is there a video of ",
        "is there a video showing ",
        "is there any clip of ",
        "is there any video of ",
        "are there any clips of ",
        "are there any videos of ",
        "did you see ",
        "do you see ",
        "have you seen ",
        "do you have ",
        "show me ",
        "find ",
        "list ",
    )

    _QUESTION_ZH_STOP_PREFIXES = ("有没有", "是否有", "是否存在", "请问有没有", "请找")

    @staticmethod
    def _strip_question_wrappers(question: str, language: str) -> str:
        text = (question or "").strip()
        if not text:
            return ""
        low = text.lower()
        if language == "zh":
            for prefix in AgentTestDatasetImporter._QUESTION_ZH_STOP_PREFIXES:
                if text.startswith(prefix):
                    text = text[len(prefix):]
                    break
            text = text.strip("？?。 ").strip()
            return text
        for prefix in AgentTestDatasetImporter._QUESTION_EN_STOP_PREFIXES:
            if low.startswith(prefix):
                text = text[len(prefix):]
                low = text.lower()
                break
        text = text.rstrip("?.! ").strip()
        if text:
            text = text[0].lower() + text[1:] if text[0].isalpha() else text
        return text

    @staticmethod
    def _build_reference_scene_description(
        *,
        recall_challenge: str | None,
        question: str,
        question_language: str,
    ) -> str | None:
        """Build the ``scene_description`` snippet that goes into ``reference_answer_rich``.

        P1-Next-F R2 (2026-05-02): the original implementation preferred the
        XLSX ``recall_challenge`` column (e.g. ``"Minimal"`` /
        ``"Must link 'other dogs approaching' to the main collision event"`` /
        ``"Bystander has no actions; easily overlooked in action-based retrieval"``).
        Those cells are *evaluator-side annotations*, not facts about the video,
        so RAGAS context_recall could never attribute them and 30/50 50-case eval
        ran with a permanently unattributable sentence in the reference.

        Now we always derive the scene snippet from the (stripped) question,
        which is a real description of what the video should depict and which
        retrieval chunks can plausibly support. ``recall_challenge`` survives
        as a pure dataset metadata field elsewhere; it just no longer leaks
        into the reference answer.

        See ``agent/recall_diagnosis_2026_05_02.md`` §2 for the diagnostic that
        motivated this change.
        """
        del recall_challenge  # intentionally unused; kept on the signature for callers
        stripped_question = AgentTestDatasetImporter._strip_question_wrappers(question, question_language)
        return stripped_question or None

    @staticmethod
    def _build_reference_answer_rich(
        *,
        video_id: str | None,
        answer_label: str,
        expected_time_raw: str | None,
        scene_description: str | None,
        question_language: str,
    ) -> str | None:
        if answer_label == "no":
            return "没有匹配片段。" if question_language == "zh" else "No matching clip is expected."
        if answer_label != "yes":
            return None
        video = (video_id or "").strip() or ("未知视频" if question_language == "zh" else "the target video")
        scene = (scene_description or "").strip()
        time_text = (expected_time_raw or "").strip()
        if question_language == "zh":
            parts: list[str] = [f"有。在视频 {video}"]
            if time_text:
                parts.append(f" 的 {time_text} 时段")
            if scene:
                parts.append(f"，{scene}")
            parts.append("。")
            return "".join(parts)
        parts_en: list[str] = [f"Yes. In {video}"]
        if time_text:
            parts_en.append(f" around {time_text}")
        if scene:
            parts_en.append(f", {scene}")
        parts_en.append(".")
        return "".join(parts_en)

    def _write_views(self, cases: list[dict[str, Any]]) -> None:
        retrieval_cases = [self._build_retrieval_view_item(case) for case in cases if case["retrieval_ready"]]
        e2e_cases = [self._build_e2e_view_item(case) for case in cases if case["e2e_ready"]]
        generation_cases = [self._build_generation_view_item(case) for case in cases if case["generation_ready"]]
        self._write_json(self.config.retrieval_view_path, retrieval_cases)
        self._write_json(self.config.e2e_view_path, e2e_cases)
        self._write_json(self.config.generation_view_path, generation_cases)

    @staticmethod
    def _build_retrieval_view_item(case: dict[str, Any]) -> dict[str, Any]:
        return {
            "case_id": case["case_id"],
            "video_id": case["video_id"],
            "query": case["question"],
            "question_language": case["question_language"],
            "recall_challenge": case["recall_challenge"],
            "expected_answer_label": case["expected_answer_label"],
            "expected_time_raw": case["expected_time_raw"],
            "expected_start_sec": case["expected_start_sec"],
            "expected_end_sec": case["expected_end_sec"],
            "difficulty_level": case["difficulty_level"],
            "video_quality_level": case["video_quality_level"],
        }

    @staticmethod
    def _build_e2e_view_item(case: dict[str, Any]) -> dict[str, Any]:
        return {
            "case_id": case["case_id"],
            "video_id": case["video_id"],
            "query": case["question"],
            "expected_answer_label": case["expected_answer_label"],
            "expected_time_raw": case["expected_time_raw"],
            "reference_answer": case["reference_answer"],
            "reference_answer_rich": case.get("reference_answer_rich"),
            "reference_scene_description": case.get("reference_scene_description"),
            "difficulty_level": case["difficulty_level"],
            "question_language": case["question_language"],
        }

    @staticmethod
    def _build_generation_view_item(case: dict[str, Any]) -> dict[str, Any]:
        return {
            "case_id": case["case_id"],
            "query": case["question"],
            "video_id": case["video_id"],
            "reference_answer": case["reference_answer"],
            "reference_answer_rich": case.get("reference_answer_rich"),
            "reference_scene_description": case.get("reference_scene_description"),
            "expected_answer_label": case["expected_answer_label"],
            "expected_time_raw": case["expected_time_raw"],
            "question_language": case["question_language"],
        }

    def _write_sqlite(self, cases: list[dict[str, Any]]) -> None:
        db_path = self.config.sqlite_path
        db_path.parent.mkdir(parents=True, exist_ok=True)
        if self.config.reset_existing and db_path.exists():
            db_path.unlink()

        with sqlite3.connect(db_path) as conn:
            conn.execute(CREATE_TABLE_SQL)
            conn.execute(f"DELETE FROM {TABLE_NAME}")
            conn.executemany(
                f"""
                INSERT INTO {TABLE_NAME} (
                    case_id, source_sheet, source_row, source_format, schema_version, video_id,
                    question, question_language, recall_challenge, expected_answer_raw,
                    expected_answer_label, expected_time_raw, expected_start_sec, expected_end_sec,
                    expected_time_is_approx, difficulty_raw, difficulty_level,
                    video_quality_raw, video_quality_level, reference_answer,
                    retrieval_ready, e2e_ready, generation_ready, metadata_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        case["case_id"],
                        case["source_sheet"],
                        case["source_row"],
                        case["source_format"],
                        case["schema_version"],
                        case["video_id"],
                        case["question"],
                        case["question_language"],
                        case["recall_challenge"],
                        case["expected_answer_raw"],
                        case["expected_answer_label"],
                        case["expected_time_raw"],
                        case["expected_start_sec"],
                        case["expected_end_sec"],
                        case["expected_time_is_approx"],
                        case["difficulty_raw"],
                        case["difficulty_level"],
                        case["video_quality_raw"],
                        case["video_quality_level"],
                        case["reference_answer"],
                        case["retrieval_ready"],
                        case["e2e_ready"],
                        case["generation_ready"],
                        json.dumps(case["metadata_json"], ensure_ascii=False),
                    )
                    for case in cases
                ],
            )
            conn.commit()

    def _build_report(self, cases: list[dict[str, Any]]) -> dict[str, Any]:
        per_sheet: dict[str, int] = {}
        answer_counts: dict[str, int] = {"yes": 0, "no": 0, "unknown": 0}
        difficulty_counts: dict[str, int] = {"easy": 0, "medium": 0, "hard": 0, "unknown": 0}
        for case in cases:
            per_sheet[case["source_sheet"]] = per_sheet.get(case["source_sheet"], 0) + 1
            answer_counts[case["expected_answer_label"]] = answer_counts.get(case["expected_answer_label"], 0) + 1
            difficulty_counts[case["difficulty_level"]] = difficulty_counts.get(case["difficulty_level"], 0) + 1
        return {
            "xlsx_path": str(self.config.xlsx_path),
            "normalized_json_path": str(self.config.normalized_json_path),
            "sqlite_path": str(self.config.sqlite_path),
            "retrieval_view_path": str(self.config.retrieval_view_path),
            "e2e_view_path": str(self.config.e2e_view_path),
            "generation_view_path": str(self.config.generation_view_path),
            "include_sheets": list(self.config.include_sheets or []),
            "total_cases": len(cases),
            "per_sheet_counts": per_sheet,
            "answer_label_counts": answer_counts,
            "difficulty_counts": difficulty_counts,
            "retrieval_ready_count": sum(case["retrieval_ready"] for case in cases),
            "e2e_ready_count": sum(case["e2e_ready"] for case in cases),
            "generation_ready_count": sum(case["generation_ready"] for case in cases),
        }

    @staticmethod
    def _write_json(path: Path, payload: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import agent_test.xlsx into normalized JSON and SQLite views")
    parser.add_argument("--xlsx-path", type=str, default=str(DEFAULT_XLSX_PATH), help="Source xlsx path")
    parser.add_argument(
        "--output-dir",
        type=str,
        default=str(DEFAULT_IMPORT_OUTPUT_DIR),
        help="Output directory for JSON/SQLite (default: agent/test/data/imported)",
    )
    parser.add_argument(
        "--sqlite-path",
        type=str,
        default=None,
        help=f"SQLite output path (default: <output-dir>/{SQLITE_EVAL_FILENAME})",
    )
    parser.add_argument("--reset", action="store_true", help="Reset existing SQLite before import")
    parser.add_argument(
        "--include-sheets",
        nargs="*",
        default=None,
        metavar="SHEET",
        help=(
            "Only parse listed worksheets (default when omitted: Part4). "
            "Part2 and Part6 are always skipped."
        ),
    )
    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help="Parse all worksheets except the hard-skipped Part2 / Part6 (overrides default Part4 filter).",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    output_dir = Path(args.output_dir).expanduser().resolve()
    if args.all_sheets:
        include_sheets: list[str] | None = None
    elif args.include_sheets is None or len(args.include_sheets) == 0:
        include_sheets = list(DEFAULT_IMPORT_INCLUDE_SHEETS)
    else:
        include_sheets = [str(item).strip() for item in args.include_sheets if str(item).strip()]
        if not include_sheets:
            include_sheets = list(DEFAULT_IMPORT_INCLUDE_SHEETS)

    sqlite_path = (
        Path(args.sqlite_path).expanduser().resolve()
        if str(args.sqlite_path or "").strip()
        else output_dir / SQLITE_EVAL_FILENAME
    )
    config = AgentTestImportConfig(
        xlsx_path=Path(args.xlsx_path).expanduser().resolve(),
        output_dir=output_dir,
        sqlite_path=sqlite_path,
        normalized_json_path=output_dir / DEFAULT_JSON_PATH.name,
        report_json_path=output_dir / DEFAULT_REPORT_PATH.name,
        retrieval_view_path=output_dir / DEFAULT_RETRIEVAL_VIEW.name,
        e2e_view_path=output_dir / DEFAULT_E2E_VIEW.name,
        generation_view_path=output_dir / DEFAULT_GENERATION_VIEW.name,
        reset_existing=bool(args.reset),
        include_sheets=include_sheets,
    )
    importer = AgentTestDatasetImporter(config)
    result = importer.build()
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
