#!/usr/bin/env python3
"""
监控管道评估：对 video_data/part4 视频跑完整流水线，并与 UCFCrime_Test.json 地面真值对齐。

流程：
  video_data/part4/{video_id}.mp4
    → Stage 1: YOLO检测+追踪+事件切片 → events.json + clips.json
    → Stage 2: LLM精炼 (vector 模式) → vector_flat.json
    → 对齐地面真值 UCFCrime_Test.json → manifest.json + summary.json

支持断点续跑：已存在的 stage1/stage2 输出自动跳过。

用法：
  python agent/test/run_pipeline_on_part4_top10.py --first-n 10    # 前 10 个
  python agent/test/run_pipeline_on_part4_top10.py --first-n 0     # 全部 52 个
  python agent/test/run_pipeline_on_part4_top10.py --first-n 0 --resume  # 断点续跑
"""

import argparse
import json
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

# 确保项目根在 sys.path
ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

# ── 路径常量 ──────────────────────────────────────────────
PART4_VIDEO_DIR = ROOT_DIR / "video_data" / "part4"
UCFCRIME_TEST_JSON = ROOT_DIR / "agent" / "test" / "data" / "UCFCrime_Test.json"
AGENT_TEST_XLSX = ROOT_DIR / "agent" / "test" / "data" / "agent_test.xlsx"
DEFAULT_OUTPUT_DIR = ROOT_DIR / "agent" / "test" / "generated" / "pipeline_eval_part4_full"


@dataclass
class PipelineConfig:
    """流水线运行参数"""
    # Stage 1 参数
    model: str = "11m"
    conf: float = 0.25
    iou: float = 0.25
    tracker: str = "botsort_reid"
    target_classes: str = "person,car,bus,truck,motorcycle,bicycle,backpack,handbag,suitcase"
    motion_threshold: float = 5.0
    min_clip_duration: float = 1.0
    max_static_duration: float = 30.0
    # Stage 2 参数
    llm_model: str = "gpt-5.4-mini"
    temperature: float = 0.1
    frames_per_sec: float = 0.1
    min_frames: int = 6
    max_frames: int = 48


@dataclass
class VideoRecord:
    """单个视频的映射关系"""
    video_id: str
    video_path: Path
    ground_truth: dict[str, Any]  # UCFCrime_Test.json 中的条目
    # 流水线输出路径（运行后填充）
    events_path: Optional[Path] = None
    clips_path: Optional[Path] = None
    vector_flat_path: Optional[Path] = None
    # 运行状态
    stage1_ok: bool = False
    stage2_ok: bool = False
    error: Optional[str] = None


def _fuzzy_match_video_id(xlsx_video_id: str, part4_files: dict[str, Path]) -> Optional[Path]:
    """将 xlsx 中的 video_id 与 part4 目录中的实际文件名做模糊匹配。

    处理两种命名差异：
      - Normal_Videos_594_x264（xlsx）↔ Normal_Videos594_x264.mp4（文件）
      - Normal_Videos_924_x264（xlsx）↔ Normal_Videos_924_x264.mp4（文件，精确匹配）
    """
    if xlsx_video_id in part4_files:
        return part4_files[xlsx_video_id]
    # 去掉 Videos_ 后面的下划线：Videos_594 → Videos594
    alt = re.sub(r"Videos_(\d)", r"Videos\1", xlsx_video_id)
    if alt in part4_files:
        return part4_files[alt]
    # 大小写不敏感 + 去下划线
    norm_xlsx = xlsx_video_id.lower().replace("_", "")
    for pf_name, pf_path in part4_files.items():
        if norm_xlsx == pf_name.lower().replace("_", ""):
            return pf_path
    return None


def load_target_video_ids_from_xlsx(
    xlsx_path: Path,
    sheet_name: str = "Part4",
) -> list[str]:
    """从 xlsx 中读取目标 sheet 的 video_id 列表（去重）。"""
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in {xlsx_path}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    video_ids: list[str] = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, [str(c) if c is not None else "" for c in row]))
        vid = row_dict.get("Video_id", "").strip()
        if vid and vid not in seen:
            seen.add(vid)
            video_ids.append(vid)
    wb.close()
    return video_ids


def discover_videos_from_xlsx(
    video_dir: Path,
    ground_truth: dict[str, Any],
    xlsx_path: Path,
    sheet_name: str = "Part4",
) -> list[VideoRecord]:
    """从 xlsx 读取目标视频列表，在 part4 目录中匹配，与 GT 对齐。

    Returns:
        VideoRecord 列表（按 xlsx 中的出现顺序）。
    """
    part4_files = {f.stem: f for f in sorted(video_dir.glob("*.mp4"))}
    if not part4_files:
        raise FileNotFoundError(f"在 {video_dir} 中未找到 .mp4 文件")

    target_ids = load_target_video_ids_from_xlsx(xlsx_path, sheet_name)
    print(f"xlsx {sheet_name} 共 {len(target_ids)} 个视频")

    records: list[VideoRecord] = []
    missing_in_part4: list[str] = []
    missing_in_gt: list[str] = []

    for xlsx_vid in target_ids:
        matched_path = _fuzzy_match_video_id(xlsx_vid, part4_files)
        if matched_path is None:
            missing_in_part4.append(xlsx_vid)
            continue

        video_id = matched_path.stem  # 使用实际文件名作为 video_id
        if video_id not in ground_truth:
            missing_in_gt.append(video_id)
            continue

        records.append(VideoRecord(
            video_id=video_id,
            video_path=matched_path,
            ground_truth=ground_truth[video_id],
        ))

    if missing_in_part4:
        print(f"  ⚠ 在 part4 中未找到: {missing_in_part4}")
    if missing_in_gt:
        print(f"  ⚠ 在 GT JSON 中未找到: {missing_in_gt}")
    print(f"  匹配成功: {len(records)} 个视频")
    return records


def run_stage1(
    record: VideoRecord,
    config: PipelineConfig,
    output_dir: Path,
) -> None:
    """Stage 1: YOLO 检测 + 追踪 + 事件切片 → events.json + clips.json"""
    from video.factory.processors.event_track_pipeline import run_pipeline as _run_pipeline
    from video.factory.processors.event_track_pipeline import save_pipeline_output

    print(f"  [Stage 1] 运行 YOLO+追踪+事件切片: {record.video_id}")

    target_classes = [x.strip() for x in config.target_classes.split(",") if x.strip()]
    start = time.time()

    events, clip_segments, meta = _run_pipeline(
        str(record.video_path),
        model_path=config.model,
        conf=config.conf,
        iou=config.iou,
        motion_threshold=config.motion_threshold,
        min_clip_duration=config.min_clip_duration,
        max_static_duration=config.max_static_duration,
        tracker=config.tracker,
        target_classes=target_classes,
    )

    # 保存文件
    events_path, clips_path = save_pipeline_output(
        events, clip_segments, meta, output_dir,
    )

    elapsed = time.time() - start
    record.events_path = events_path
    record.clips_path = clips_path
    record.stage1_ok = True

    print(f"    ✓ 完成 ({elapsed:.1f}s): {len(events)} 事件, {len(clip_segments)} 片段")
    print(f"      events: {events_path}")
    print(f"      clips:  {clips_path}")


def run_stage2(
    record: VideoRecord,
    config: PipelineConfig,
    output_dir: Path,
) -> None:
    """Stage 2: LLM 精炼 (vector 模式) → vector_flat.json"""
    from dotenv import load_dotenv
    load_dotenv(ROOT_DIR / ".env")

    from video.factory.refinement_runner import RefineEventsConfig, run_refine_events_from_files

    print(f"  [Stage 2] LLM 精炼 (vector 模式): {record.video_id}")

    assert record.events_path and record.clips_path, "先运行 Stage 1"

    cfg = RefineEventsConfig(
        mode="vector",
        model=config.llm_model,
        temperature=config.temperature,
        frames_per_sec=config.frames_per_sec,
        min_frames=config.min_frames,
        max_frames=config.max_frames,
    )

    start = time.time()
    out_path = run_refine_events_from_files(
        record.events_path, record.clips_path, cfg,
    )
    elapsed = time.time() - start

    record.vector_flat_path = out_path
    record.stage2_ok = True

    # 读取输出统计
    out_data = json.loads(out_path.read_text(encoding="utf-8"))
    event_count = len(out_data.get("events", []))
    print(f"    ✓ 完成 ({elapsed:.1f}s): {event_count} 精炼事件")
    print(f"      vector_flat: {out_path}")


def save_ground_truth(record: VideoRecord, output_dir: Path) -> None:
    """将 ground truth 单独保存一份到输出目录，方便后续评估对比"""
    gt_path = output_dir / f"{record.video_id}_ground_truth.json"
    gt_path.write_text(
        json.dumps(record.ground_truth, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def build_manifest(
    records: list[VideoRecord],
    output_dir: Path,
    config: PipelineConfig,
    elapsed_total: float,
) -> dict[str, Any]:
    """生成 manifest.json"""
    items = []
    for r in records:
        item = {
            "video_id": r.video_id,
            "video_path": str(r.video_path),
            "ground_truth_duration": r.ground_truth.get("duration"),
            "ground_truth_sentences": len(r.ground_truth.get("sentences", [])),
            "stage1_ok": r.stage1_ok,
            "stage2_ok": r.stage2_ok,
            "events_path": str(r.events_path) if r.events_path else None,
            "clips_path": str(r.clips_path) if r.clips_path else None,
            "vector_flat_path": str(r.vector_flat_path) if r.vector_flat_path else None,
            "error": r.error,
        }
        items.append(item)

    return {
        "pipeline_config": {
            "model": config.model,
            "conf": config.conf,
            "iou": config.iou,
            "tracker": config.tracker,
            "target_classes": config.target_classes,
            "llm_model": config.llm_model,
            "temperature": config.temperature,
        },
        "output_dir": str(output_dir),
        "video_count": len(records),
        "elapsed_total_sec": round(elapsed_total, 1),
        "items": items,
    }


def build_summary(records: list[VideoRecord]) -> dict[str, Any]:
    """生成 summary.json"""
    total = len(records)
    stage1_ok = sum(1 for r in records if r.stage1_ok)
    stage2_ok = sum(1 for r in records if r.stage2_ok)
    failed = sum(1 for r in records if r.error)

    # 汇总 ground truth 统计
    total_gt_sentences = sum(len(r.ground_truth.get("sentences", [])) for r in records)
    total_gt_duration = sum(r.ground_truth.get("duration", 0) or 0 for r in records)

    # 汇总 pipeline 输出
    total_events = 0
    total_refined_events = 0
    for r in records:
        if r.events_path and r.events_path.exists():
            ev = json.loads(r.events_path.read_text(encoding="utf-8"))
            total_events += len(ev.get("events", []))
        if r.vector_flat_path and r.vector_flat_path.exists():
            vf = json.loads(r.vector_flat_path.read_text(encoding="utf-8"))
            total_refined_events += len(vf.get("events", []))

    return {
        "total_videos": total,
        "stage1_completed": stage1_ok,
        "stage2_completed": stage2_ok,
        "failed": failed,
        "ground_truth": {
            "total_sentences": total_gt_sentences,
            "total_duration_sec": round(total_gt_duration, 1),
        },
        "pipeline_output": {
            "total_raw_events": total_events,
            "total_refined_events": total_refined_events,
        },
    }


def print_summary_table(records: list[VideoRecord]) -> None:
    """打印简洁的汇总表"""
    print("\n" + "=" * 80)
    print(f"{'#':>3}  {'Video ID':<35} {'GT':>5} {'S1':>4} {'S2':>4}  {'Error'}")
    print("-" * 80)
    for i, r in enumerate(records, 1):
        gt_n = len(r.ground_truth.get("sentences", []))
        s1 = "✓" if r.stage1_ok else "✗"
        s2 = "✓" if r.stage2_ok else "✗"
        err = r.error or ""
        if len(err) > 30:
            err = err[:27] + "..."
        print(f"{i:>3}  {r.video_id:<35} {gt_n:>5} {s1:>4} {s2:>4}  {err}")
    print("-" * 80)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="监控管道评估：part4 视频 → 流水线 → 对齐地面真值",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--first-n", type=int, default=10,
                   help="跑前 N 个视频，0=全部 (default: 10)")
    p.add_argument("--from-xlsx", action="store_true",
                   help="从 agent_test.xlsx 的 Part4 sheet 读取目标视频列表（覆盖 --first-n）")
    p.add_argument("--output-dir", type=str, default=str(DEFAULT_OUTPUT_DIR),
                   help="输出目录 (default: agent/test/generated/pipeline_eval_part4_full/)")
    p.add_argument("--video-dir", type=str, default=str(PART4_VIDEO_DIR),
                   help="视频源目录")
    p.add_argument("--ground-truth", type=str, default=str(UCFCRIME_TEST_JSON),
                   help="地面真值 JSON 路径")
    p.add_argument("--skip-stage1", action="store_true",
                   help="跳过 Stage 1（使用已有 events/clips.json）")
    p.add_argument("--skip-stage2", action="store_true",
                   help="跳过 Stage 2（使用已有 vector_flat.json）")
    p.add_argument("--resume", action="store_true",
                   help="断点续跑：自动跳过已完成的 Stage")
    p.add_argument("--model", type=str, default="11m",
                   help="YOLO 模型 (default: 11m)")
    p.add_argument("--conf", type=float, default=0.25,
                   help="检测置信度阈值 (default: 0.25)")
    p.add_argument("--llm-model", type=str, default="gpt-5.4-mini",
                   help="LLM 模型 (default: gpt-5.4-mini)")
    return p


def main() -> None:
    args = build_parser().parse_args()

    # ── 初始化 ────────────────────────────────────────────
    output_dir = Path(args.output_dir).resolve()
    video_dir = Path(args.video_dir).resolve()
    ground_truth_path = Path(args.ground_truth).resolve()

    if not video_dir.is_dir():
        print(f"错误: 视频目录不存在: {video_dir}")
        sys.exit(1)
    if not ground_truth_path.exists():
        print(f"错误: 地面真值文件不存在: {ground_truth_path}")
        sys.exit(1)

    # 子目录
    stage1_dir = output_dir / "stage1"
    stage2_dir = output_dir / "stage2"
    gt_dir = output_dir / "ground_truth"
    for d in [stage1_dir, stage2_dir, gt_dir]:
        d.mkdir(parents=True, exist_ok=True)

    config = PipelineConfig(
        model=args.model,
        conf=args.conf,
        llm_model=args.llm_model,
    )

    # ── 发现视频 & 对齐地面真值 ──────────────────────────
    print("=" * 80)
    print("监控管道评估：part4 → 流水线 → 对齐地面真值")
    print("=" * 80)
    print(f"视频目录:   {video_dir}")
    print(f"地面真值:   {ground_truth_path}")
    print(f"输出目录:   {output_dir}")
    print(f"取前 N:     {'全部' if args.first_n <= 0 else args.first_n}")
    print(f"从xlsx读取: {args.from_xlsx}")
    print(f"跳过Stage1: {args.skip_stage1}")
    print(f"跳过Stage2: {args.skip_stage2}")
    print(f"断点续跑:   {args.resume}")
    print()

    ground_truth = json.loads(ground_truth_path.read_text(encoding="utf-8"))
    print(f"地面真值共 {len(ground_truth)} 个视频条目\n")

    records = discover_videos_from_xlsx(
        video_dir, ground_truth,
        xlsx_path=Path(args.xlsx_path) if hasattr(args, 'xlsx_path') else AGENT_TEST_XLSX,
    )
    if not records:
        print("没有找到可匹配的视频，退出。")
        sys.exit(1)

    print(f"\nxlsx 目标视频列表：")
    for i, r in enumerate(records, 1):
        gt = r.ground_truth
        print(f"  {i:>2}. {r.video_id}  ({gt.get('duration', '?'):.1f}s, {len(gt.get('sentences', []))} 句描述)")
    print()

    # ── Stage 1: YOLO + 追踪 + 事件切片 ─────────────────
    total_start = time.time()

    def _check_stage1_done(r: VideoRecord) -> bool:
        ev_p = stage1_dir / f"{r.video_id}_events.json"
        cl_p = stage1_dir / f"{r.video_id}_clips.json"
        if ev_p.exists() and cl_p.exists():
            r.events_path = ev_p
            r.clips_path = cl_p
            r.stage1_ok = True
            return True
        return False

    def _check_stage2_done(r: VideoRecord) -> bool:
        vf_p = stage2_dir / f"{r.video_id}_events_vector_flat.json"
        if vf_p.exists():
            r.vector_flat_path = vf_p
            r.stage2_ok = True
            return True
        return False

    # 断点续跑：先检查已完成的
    if args.resume:
        s1_skipped = sum(1 for r in records if _check_stage1_done(r))
        s2_skipped = sum(1 for r in records if _check_stage2_done(r))
        if s1_skipped or s2_skipped:
            print(f"[断点续跑] Stage1 已完成 {s1_skipped} 个, Stage2 已完成 {s2_skipped} 个")

    if not args.skip_stage1:
        print("━" * 80)
        print("Stage 1: YOLO + BoT-SORT 追踪 + 事件切片")
        print("━" * 80)
        s1_pending = [r for r in records if not r.stage1_ok]
        if not s1_pending:
            print("  全部已完成，跳过。")
        for i, record in enumerate(s1_pending, 1):
            print(f"\n[{i}/{len(s1_pending)}] {record.video_id}")
            try:
                run_stage1(record, config, stage1_dir)
            except Exception as e:
                record.error = f"Stage1: {e}"
                print(f"    ✗ 失败: {e}")
    else:
        print("跳过 Stage 1，尝试读取已有的 events.json / clips.json")
        for r in records:
            ev_p = stage1_dir / f"{r.video_id}_events.json"
            cl_p = stage1_dir / f"{r.video_id}_clips.json"
            if ev_p.exists() and cl_p.exists():
                r.events_path = ev_p
                r.clips_path = cl_p
                r.stage1_ok = True
                print(f"  ✓ {r.video_id}: 已有 Stage 1 输出")
            else:
                r.error = "Stage1: 缺少已有输出"
                print(f"  ✗ {r.video_id}: 缺少 Stage 1 输出，但 --skip-stage1 已指定")

    # ── Stage 2: LLM 精炼 ─────────────────────────────────
    if not args.skip_stage2:
        print("\n" + "━" * 80)
        print("Stage 2: LLM 精炼 (vector 模式)")
        print("━" * 80)
        s2_pending = [r for r in records if r.stage1_ok and not r.stage2_ok]
        if not s2_pending:
            print("  全部已完成，跳过。")
        for i, record in enumerate(s2_pending, 1):
            print(f"\n[{i}/{len(s2_pending)}] {record.video_id}")
            try:
                run_stage2(record, config, stage2_dir)
            except Exception as e:
                record.error = (record.error or "") + f"; Stage2: {e}"
                print(f"    ✗ 失败: {e}")
    else:
        print("跳过 Stage 2，尝试读取已有的 vector_flat.json")
        for r in records:
            vf_p = stage2_dir / f"{r.video_id}_events_vector_flat.json"
            if vf_p.exists():
                r.vector_flat_path = vf_p
                r.stage2_ok = True
                print(f"  ✓ {r.video_id}: 已有 Stage 2 输出")
            elif r.stage1_ok:
                r.error = (r.error or "") + "; Stage2: 缺少已有输出"
                print(f"  ✗ {r.video_id}: 缺少 Stage 2 输出，但 --skip-stage2 已指定")

    # ── 保存地面真值副本 ──────────────────────────────────
    print("\n" + "━" * 80)
    print("保存地面真值副本")
    print("━" * 80)
    for r in records:
        save_ground_truth(r, gt_dir)
        print(f"  ✓ {r.video_id}_ground_truth.json")

    # ── 生成 manifest & summary ───────────────────────────
    elapsed_total = time.time() - total_start

    manifest = build_manifest(records, output_dir, config, elapsed_total)
    manifest_path = output_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    summary = build_summary(records)
    summary["elapsed_total_sec"] = round(elapsed_total, 1)
    summary_path = output_dir / "summary.json"
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # ── 打印汇总 ──────────────────────────────────────────
    print("\n" + "=" * 80)
    print("完成汇总")
    print("=" * 80)
    print(f"总耗时: {elapsed_total:.1f}s")
    print(f"Stage 1 成功: {summary['stage1_completed']}/{summary['total_videos']}")
    print(f"Stage 2 成功: {summary['stage2_completed']}/{summary['total_videos']}")
    print(f"地面真值句子总数: {summary['ground_truth']['total_sentences']}")
    print(f"地面真值总时长: {summary['ground_truth']['total_duration_sec']:.1f}s")
    print(f"原始事件总数: {summary['pipeline_output']['total_raw_events']}")
    print(f"精炼事件总数: {summary['pipeline_output']['total_refined_events']}")
    print(f"\nManifest: {manifest_path}")
    print(f"Summary:  {summary_path}")

    print_summary_table(records)

    print(f"\n下一步：将 {stage2_dir} 下的 *_vector_flat.json 用于评估比对")


if __name__ == "__main__":
    main()
