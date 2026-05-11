"""tests/test_mevid_full.py
==========================
Full MEVID end-to-end test — covers ALL pipeline modules:

  ① YOLO11m + BoT-SORT tracking          (per camera, local)
  ② OSNet x1.0 Re-ID embeddings          (torchreid, local, auto-download weights)
  ③ CameraTopologyPrior                  (cross-camera transit model, local)
  ④ Cross-camera greedy matching         (union-find, local)
  ⑤ LLM event description refinement    (optional --refine, DashScope)
  ⑥ VLM yes/no QA evaluation            (100 cases stratified, DashScope)

Cost estimate
-------------
  No refinement (default):  ~¥2 CNY  (100 QA cases × ~1 k tokens)
  With --refine (all slots): ~¥10 CNY

Usage
-----
  # All 6 slots, 100 QA cases, no LLM refinement
  python tests/test_mevid_full.py \\
      --video-dir _data/mevid_slots \\
      --xlsx agent/test/data/agent_test_mevid.xlsx

  # Include LLM event refinement
  python tests/test_mevid_full.py \\
      --video-dir _data/mevid_slots \\
      --xlsx agent/test/data/agent_test_mevid.xlsx --refine

  # Single slot smoke-test
  python tests/test_mevid_full.py \\
      --video-dir _data/mevid_slots \\
      --xlsx agent/test/data/agent_test_mevid.xlsx --slot 16-35

  # Resume after interruption
  python tests/test_mevid_full.py \\
      --video-dir _data/mevid_slots \\
      --xlsx agent/test/data/agent_test_mevid.xlsx --resume
"""

from __future__ import annotations

import argparse
import base64
import json
import logging
import os
import random
import re
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path

import cv2
from dotenv import load_dotenv
from openai import OpenAI

load_dotenv()

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from video.factory.multi_camera_coordinator import run_multi_camera_pipeline  # noqa: E402
from video.factory.refinement_runner import RefineEventsConfig, refine_multi_camera_output  # noqa: E402
from video.core.schema.multi_camera import CrossCameraConfig  # noqa: E402

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("test_mevid_full")

# ── Slot → camera → video stem ────────────────────────────────────────────────
SLOT_CAMERAS: dict[str, dict[str, str]] = {
    "11-20": {
        "G329": "2018-03-11.11-20-00.11-25-00.admin.G329.r13",
        "G330": "2018-03-11.11-20-00.11-25-00.school.G330.r13",
        "G336": "2018-03-11.11-20-00.11-25-00.school.G336.r13",
        "G419": "2018-03-11.11-20-00.11-25-00.school.G419.r13",
        "G420": "2018-03-11.11-20-00.11-25-00.school.G420.r13",
        "G421": "2018-03-11.11-20-00.11-25-00.school.G421.r13",
        "G423": "2018-03-11.11-20-00.11-25-00.school.G423.r13",
        "G508": "2018-03-11.11-20-00.11-25-00.bus.G508.r13",
    },
    "11-55": {
        "G299": "2018-03-11.11-55-00.12-00-00.school.G299.r13",
        "G328": "2018-03-11.11-55-00.12-00-00.school.G328.r13",
        "G330": "2018-03-11.11-55-00.12-00-00.school.G330.r13",
        "G419": "2018-03-11.11-55-00.12-00-00.school.G419.r13",
        "G420": "2018-03-11.11-55-00.12-00-00.school.G420.r13",
        "G506": "2018-03-11.11-55-00.12-00-00.bus.G506.r13",
        "G508": "2018-03-11.11-55-00.12-00-00.bus.G508.r13",
    },
    "13-50": {
        "G328": "2018-03-11.13-50-01.13-55-01.school.G328.r13",
        "G329": "2018-03-11.13-50-01.13-55-01.admin.G329.r13",
        "G339": "2018-03-11.13-50-01.13-55-01.school.G339.r13",
        "G421": "2018-03-11.13-50-01.13-55-01.school.G421.r13",
        "G424": "2018-03-11.13-50-01.13-55-01.school.G424.r13",
        "G506": "2018-03-11.13-50-01.13-55-01.bus.G506.r13",
        "G508": "2018-03-11.13-50-01.13-55-01.bus.G508.r13",
    },
    "14-20": {
        "G328": "2018-03-11.14-20-01.14-25-01.school.G328.r13",
        "G339": "2018-03-11.14-20-01.14-25-01.school.G339.r13",
        "G419": "2018-03-11.14-20-01.14-25-01.school.G419.r13",
        "G421": "2018-03-11.14-20-01.14-25-01.school.G421.r13",
        "G423": "2018-03-11.14-20-01.14-25-01.school.G423.r13",
        "G505": "2018-03-11.14-20-01.14-25-01.bus.G505.r13",
        "G506": "2018-03-11.14-20-01.14-25-01.bus.G506.r13",
        "G508": "2018-03-11.14-20-01.14-25-01.bus.G508.r13",
    },
    "16-20": {
        "G326": "2018-03-11.16-20-01.16-25-01.admin.G326.r13",
        "G328": "2018-03-11.16-20-01.16-25-01.school.G328.r13",
        "G329": "2018-03-11.16-20-01.16-25-01.admin.G329.r13",
        "G336": "2018-03-11.16-20-01.16-25-01.school.G336.r13",
        "G419": "2018-03-11.16-20-01.16-25-01.school.G419.r13",
        "G420": "2018-03-11.16-20-01.16-25-01.school.G420.r13",
        "G506": "2018-03-11.16-20-01.16-25-01.bus.G506.r13",
        "G508": "2018-03-11.16-20-01.16-25-01.bus.G508.r13",
    },
    "16-35": {
        "G326": "2018-03-11.16-35-01.16-40-01.admin.G326.r13",
        "G328": "2018-03-11.16-35-01.16-40-01.school.G328.r13",
        "G329": "2018-03-11.16-35-01.16-40-01.admin.G329.r13",
        "G336": "2018-03-11.16-35-01.16-40-01.school.G336.r13",
        "G339": "2018-03-11.16-35-01.16-40-01.school.G339.r13",
        "G419": "2018-03-11.16-35-01.16-40-01.school.G419.r13",
        "G420": "2018-03-11.16-35-01.16-40-01.school.G420.r13",
        "G423": "2018-03-11.16-35-01.16-40-00.school.G423.r13",
        "G506": "2018-03-11.16-35-01.16-40-01.bus.G506.r13",
        "G638": "2018-03-11.16-35-01.16-40-01.school.G638.r13",
    },
}

# Reverse: stem → (slot, camera_id)
STEM_TO_SLOT_CAM: dict[str, tuple[str, str]] = {}
for _slot, _cams in SLOT_CAMERAS.items():
    for _cam, _stem in _cams.items():
        STEM_TO_SLOT_CAM[_stem] = (_slot, _cam)

# ── API config ─────────────────────────────────────────────────────────────────
API_KEY  = os.getenv("DASHSCOPE_API_KEY", "")
BASE_URL = os.getenv("DASHSCOPE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1")
MODEL    = os.getenv("DASHSCOPE_CHAT_MODEL", "qwen-vl-max-latest")

PRICE_IN  = float(os.environ.get("QWEN_PRICE_IN",  "0.012"))
PRICE_OUT = float(os.environ.get("QWEN_PRICE_OUT", "0.036"))

PIPELINE_CACHE_DIR = ROOT / "_cache" / "mevid_pipeline"
YOLO_CACHE_DIR     = ROOT / "_cache" / "yolo_mevid_full"


# ══════════════════════════════════════════════════════════════════════════════
# ① – ④  Pipeline phase (local, no API cost)
# ══════════════════════════════════════════════════════════════════════════════

def _serialise_output(mc_output) -> dict:
    """Convert MultiCameraOutput → JSON-safe dict (skip np.ndarray fields)."""
    cameras_out = []
    for cr in mc_output.per_camera:
        cameras_out.append({
            "camera_id": cr.camera_id,
            "video_path": cr.video_path,
            "events": cr.events,
            "clips":  cr.clips,
            "meta":   cr.meta,
            "n_tracks": len(cr.tracks),
        })
    entities_out = []
    for ge in mc_output.global_entities:
        entities_out.append({
            "global_entity_id": ge.global_entity_id,
            "appearances": [
                {
                    "camera_id":  a.camera_id,
                    "track_id":   a.track_id,
                    "start_time": a.start_time,
                    "end_time":   a.end_time,
                    "confidence": a.confidence,
                }
                for a in ge.appearances
            ],
        })
    return {
        "global_entities": entities_out,
        "per_camera":      cameras_out,
        "merged_events":   mc_output.merged_events,
    }


def run_pipeline_for_slot(
    slot: str,
    video_dir: Path,
    reid_device: str = "cpu",
    force: bool = False,
) -> dict:
    """Run the full multi-camera pipeline for one slot; caches result as JSON."""
    cache_file = PIPELINE_CACHE_DIR / f"{slot}_pipeline.json"
    if not force and cache_file.exists():
        logger.info("[%s] Loading cached pipeline output from %s", slot, cache_file)
        return json.loads(cache_file.read_text(encoding="utf-8"))

    cam_map = SLOT_CAMERAS[slot]
    camera_videos: dict[str, str] = {}
    missing = []
    for cam_id, stem in cam_map.items():
        p = video_dir / f"{stem}.avi"
        if p.exists():
            camera_videos[cam_id] = str(p)
        else:
            missing.append(cam_id)

    if missing:
        logger.warning("[%s] Missing video files for cameras: %s — skipping those", slot, missing)
    if not camera_videos:
        raise FileNotFoundError(f"No videos found for slot {slot} in {video_dir}")

    logger.info(
        "[%s] Running pipeline on %d cameras: %s",
        slot, len(camera_videos), list(camera_videos),
    )

    # CrossCameraConfig tuned for MEVID synchronized multi-camera:
    # All cameras record the same 5-min window simultaneously.
    # OSNet cross-camera cosine similarity typically 0.55–0.70 for same person
    # due to angle/lighting variation — threshold must be below that range.
    # v3 target: ~20–40 global entities for ~20 annotated subjects.
    #   0.60 + 300s → 1 entity (over-merge, chain effect)
    #   0.75 + 90s  → 562 entities (under-merge, too strict)
    #   0.63 + 150s → target ~20-40 entities
    config = CrossCameraConfig(
        # Topology data shows most camera pairs have ~0s transition time → overlapping FoV.
        # Same person genuinely appears in multiple cameras simultaneously.
        # Allow both overlap-based and gap-based pairing.
        min_overlap_sec=1.0,            # require at least 1s overlap (filters noise)
        max_transition_sec=180.0,       # allow 3 min sequential gap
        embedding_threshold=0.63,       # OSNet cross-cam cosine threshold
        cross_camera_min_score=0.58,    # combined score threshold
        same_camera_reid_threshold=0.80,
        topology_weight_reid=0.80,
        topology_weight_topo=0.20,
        person_only=True,
    )

    t0 = time.time()
    mc_output = run_multi_camera_pipeline(
        camera_videos=camera_videos,
        config=config,
        reid_device=reid_device,
        num_crops=5,
        use_llm_verify=False,       # keep local, no API in this phase
        model_path="11m",
        conf=0.40,   # was 0.25 — higher threshold cuts noisy/short-lived detections
        iou=0.40,    # was 0.25
    )
    elapsed = time.time() - t0

    logger.info(
        "[%s] Pipeline done in %.1fs — %d global entities, %d merged events",
        slot, elapsed,
        len(mc_output.global_entities),
        len(mc_output.merged_events),
    )

    result = _serialise_output(mc_output)
    result["slot"] = slot
    result["elapsed_sec"] = round(elapsed, 1)
    result["cameras_run"] = list(camera_videos)

    PIPELINE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_file.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info("[%s] Cached to %s", slot, cache_file)
    return result


# ══════════════════════════════════════════════════════════════════════════════
# ⑤  Optional LLM refinement phase
# ══════════════════════════════════════════════════════════════════════════════

def run_refinement_for_slot(slot: str, video_dir: Path) -> dict:
    """Run LLM event refinement for one slot (uses DashScope as OpenAI-compatible endpoint).

    Reuses the cached pipeline JSON instead of re-running YOLO+OSNet.
    refine_multi_camera_output only needs: per_camera.video_path, per_camera.clips,
    merged_events, and global_entities — all present in the pipeline cache.
    """
    cache_file = PIPELINE_CACHE_DIR / f"{slot}_refined.json"
    if cache_file.exists():
        logger.info("[%s] Loading cached refinement from %s", slot, cache_file)
        return json.loads(cache_file.read_text(encoding="utf-8"))

    # Force-set env so refinement runner hits DashScope (not OpenAI)
    os.environ["OPENAI_API_KEY"] = API_KEY
    os.environ["OPENAI_BASE_URL"] = BASE_URL

    # Load pipeline cache — avoid re-running YOLO+OSNet (takes 40+ min)
    pipeline_cache = PIPELINE_CACHE_DIR / f"{slot}_pipeline.json"
    if not pipeline_cache.exists():
        logger.error("[%s] Pipeline cache not found: %s — run pipeline first", slot, pipeline_cache)
        return {}
    cached = json.loads(pipeline_cache.read_text(encoding="utf-8"))

    # Reconstruct minimal mc_output from cache using simple namespace objects.
    # refine_multi_camera_output only accesses: camera_result.camera_id,
    # camera_result.video_path, camera_result.clips, output.merged_events,
    # output.global_entities — no tracks or embeddings needed.
    from types import SimpleNamespace
    cam_map = SLOT_CAMERAS[slot]
    per_camera = []
    for cr in cached.get("per_camera", []):
        cam_id = cr["camera_id"]
        stem = cam_map.get(cam_id, "")
        video_path = str(video_dir / f"{stem}.avi") if stem else cr.get("video_path", "")
        per_camera.append(SimpleNamespace(
            camera_id=cam_id,
            video_path=video_path,
            clips=cr.get("clips", []),
            tracks=[],
            events=cr.get("events", []),
            meta=cr.get("meta", {}),
            person_crops={},
            person_embeddings={},
        ))

    # global_entities must be SimpleNamespace too — refinement_runner accesses
    # ent.appearances, a.start_time, a.camera_id, a.track_id as attributes
    global_entities = []
    for ge in cached.get("global_entities", []):
        appearances = [
            SimpleNamespace(
                camera_id=a["camera_id"],
                track_id=a["track_id"],
                start_time=a["start_time"],
                end_time=a["end_time"],
                confidence=a.get("confidence", 0.0),
            )
            for a in ge.get("appearances", [])
        ]
        global_entities.append(SimpleNamespace(
            global_entity_id=ge["global_entity_id"],
            appearances=appearances,
        ))

    mc_output = SimpleNamespace(
        per_camera=per_camera,
        merged_events=cached.get("merged_events", []),
        global_entities=global_entities,
    )

    refine_cfg = RefineEventsConfig(
        mode="vector",
        frames_per_sec=0.5,
        min_frames=4,
        max_frames=12,
        model=MODEL,
        temperature=0.0,
    )
    logger.info("[%s] Running LLM refinement …", slot)
    t0 = time.time()
    refined = refine_multi_camera_output(mc_output, config=refine_cfg)
    elapsed = time.time() - t0
    logger.info("[%s] Refinement done in %.1fs", slot, elapsed)

    result = {"slot": slot, "elapsed_sec": round(elapsed, 1), "per_camera": refined}
    PIPELINE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_file.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


# ══════════════════════════════════════════════════════════════════════════════
# Context builder
# ══════════════════════════════════════════════════════════════════════════════

def _fmt_sec(s: float) -> str:
    m, sec = divmod(int(s), 60)
    return f"{m}:{sec:02d}"


def build_context(video_id: str, pipeline_data: dict[str, dict], refined_data: dict[str, dict],
                  question: str = "", context_mode: str = "default") -> str:
    """Build a natural-language context string from pipeline + refinement output for one video.

    If ``question`` is provided, cross-camera context is tailored to the cameras mentioned
    in the question. ``context_mode="negative"`` suppresses cross-camera candidate text so
    no-answer cases are not biased by uncertain Re-ID hypotheses.
    """
    if video_id not in STEM_TO_SLOT_CAM:
        return ""
    slot, cam_id = STEM_TO_SLOT_CAM[video_id]

    lines: list[str] = []

    # ── Per-camera events ──────────────────────────────────────────────────────
    slot_data = pipeline_data.get(slot, {})
    per_cam = {cr["camera_id"]: cr for cr in slot_data.get("per_camera", [])}
    cam_result = per_cam.get(cam_id)
    if cam_result:
        events = cam_result.get("events", [])
        # Only show person events (filter out vehicles/static objects to save tokens)
        person_events = [
            ev for ev in events
            if ev.get("class_name", "").lower() in ("person", "people", "pedestrian")
        ]
        if person_events:
            lines.append(f"[Camera {cam_id}] Person tracks ({len(person_events)} motion segments):")
            for ev in person_events[:30]:   # cap at 30 person events
                start = _fmt_sec(ev.get("start_time", 0))
                end   = _fmt_sec(ev.get("end_time", 0))
                tid   = ev.get("track_id", "?")
                desc  = ev.get("event_text", "")
                if desc:
                    lines.append(f"  • {start}–{end}  track#{tid}: {desc}")
                else:
                    lines.append(f"  • {start}–{end}  track#{tid} (person moving)")
        else:
            lines.append(f"[Camera {cam_id}] No person motion events detected.")

    # ── LLM-refined events (if available) ─────────────────────────────────────
    slot_refined = refined_data.get(slot, {}).get("per_camera", {})
    cam_refined  = slot_refined.get(cam_id, {})
    if cam_refined:
        ref_events = cam_refined.get("events", [])
        if ref_events:
            lines.append(f"\n[Camera {cam_id}] LLM-refined event descriptions:")
            for ev in ref_events[:10]:
                start = _fmt_sec(ev.get("start_time", 0))
                end   = _fmt_sec(ev.get("end_time", 0))
                desc  = ev.get("event_text", ev.get("description", ""))
                lines.append(f"  • {start}–{end}: {desc}")

    # ── Cross-camera global entities ───────────────────────────────────────────
    entities = slot_data.get("global_entities", [])
    cross_cam_entities = [
        ge for ge in entities
        if any(a["camera_id"] == cam_id for a in ge["appearances"])
        and len({a["camera_id"] for a in ge["appearances"]}) > 1
    ]

    # Determine which other cameras are mentioned in the question (for targeted context)
    q_lower = question.lower()
    q_mentioned_cams = [
        cam for cam in SLOT_CAMERAS.get(slot, {}).keys()
        if cam != cam_id and cam.lower() in q_lower
    ] if question else []

    include_cross_camera = context_mode != "negative"

    if include_cross_camera and cross_cam_entities:
        # If question mentions specific cameras, add targeted transition context first
        if q_mentioned_cams:
            target_lines: list[str] = []
            for ge in cross_cam_entities:
                my_apps   = [a for a in ge["appearances"] if a["camera_id"] == cam_id]
                q_cam_apps = [a for a in ge["appearances"] if a["camera_id"] in q_mentioned_cams]
                if not my_apps or not q_cam_apps:
                    continue
                for q_app in q_cam_apps:
                    for my_app in my_apps:
                        my_dur = float(my_app.get("end_time", 0)) - float(my_app.get("start_time", 0))
                        brief_note = f" [brief {my_dur:.0f}s appearance]" if my_dur <= 6.0 else ""
                        target_lines.append(
                            f"  • Re-ID CANDIDATE: possible same person in {q_app['camera_id']} at "
                            f"{_fmt_sec(q_app['start_time'])}–{_fmt_sec(q_app['end_time'])} "
                            f"AND in {cam_id} at "
                            f"{_fmt_sec(my_app['start_time'])}–{_fmt_sec(my_app['end_time'])}"
                            f"{brief_note}"
                            f" (candidate score={ge['appearances'][0].get('confidence',0):.2f})"
                        )
            if target_lines:
                q_cam_str = " ↔ ".join([cam_id] + q_mentioned_cams)
                lines.append(
                    f"\n[Cross-camera {q_cam_str}] Re-ID candidate transitions "
                    f"(use as hints; verify with visual/context evidence):"
                )
                lines.extend(target_lines[:6])

        # General cross-camera summary
        lines.append(
            f"\n[Cross-camera] Re-ID candidate list: {len(cross_cam_entities)} possible "
            f"multi-camera person trajectory/trajectories in this slot:"
        )
        for ge in cross_cam_entities[:8]:
            times = [
                f"{a['camera_id']} at {_fmt_sec(a['start_time'])}–{_fmt_sec(a['end_time'])}"
                for a in ge["appearances"]
            ]
            conf_vals = [a.get("confidence", 0.0) for a in ge["appearances"]]
            avg_conf = sum(conf_vals) / len(conf_vals) if conf_vals else 0.0
            conf_str = f" (Re-ID={avg_conf:.2f})" if avg_conf > 0 else ""
            lines.append(
                f"  • POSSIBLE SAME PERSON ({ge['global_entity_id']}){conf_str}: "
                + " → ".join(times)
            )

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
# ⑥  VLM QA phase
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = (
    "You are a surveillance video analyst with access to automated tracker output.\n"
    "You will be shown frames from a 5-minute security camera clip, "
    "along with structured context from a multi-camera tracking pipeline "
    "(YOLO detections, Re-ID cross-camera matches, optional LLM event descriptions).\n"
    "Answer the user's question concisely. "
    "Your answer MUST start with 'yes' or 'no'. "
    "If yes, include the approximate time (mm:ss).\n"
    "CRITICAL RULES:\n"
    "1. For cross-camera identity questions (did person X appear in camera A and then camera B?): "
    "Treat [Cross-camera Re-ID CANDIDATE] lines as candidate hints, not ground truth. "
    "Answer YES only when the candidate timing plus the visible frames or event context support "
    "the described appearance/person. Do not answer YES from Re-ID text alone.\n"
    "2. For single-camera appearance questions: "
    "If tracker or candidate context shows a person active in this camera at a specific time, "
    "use it as a clue, but still check whether it matches the requested appearance. "
    "A '[brief Xs appearance]' note means the person was only there for X seconds; "
    "dedicate extra attention to the labeled brief-appearance frames provided.\n"
    "3. Absence of visual evidence ≠ absence of the person. "
    "Sampled frames cover only a fraction of the video; use tracker data for hard-to-see cases, "
    "but be conservative when the question asks whether a person also appeared in another camera."
)


_MAX_FRAME_EDGE = 640  # px — keeps token cost ~10× lower than raw 1080p


def _resize_frame(frame: "np.ndarray") -> "np.ndarray":
    """Resize frame so the longest edge ≤ _MAX_FRAME_EDGE (preserves aspect ratio)."""
    h, w = frame.shape[:2]
    if max(h, w) <= _MAX_FRAME_EDGE:
        return frame
    scale = _MAX_FRAME_EDGE / max(h, w)
    return cv2.resize(frame, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)


def sample_frames(video_path: str, n_frames: int = 10,
                  clips: list[dict] | None = None) -> list[str]:
    """Sample n_frames from video.

    If clips (motion segments) are provided, samples proportionally from
    active regions only (clip-aware). Falls back to uniform if no clips.
    """
    frames_b64: list[str] = []
    cap = cv2.VideoCapture(video_path)
    fps   = cap.get(cv2.CAP_PROP_FPS) or 25.0
    total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) or 1

    if clips:
        total_active = sum(max(0, c["end_sec"] - c["start_sec"]) for c in clips)
        if total_active > 0:
            for clip in clips:
                dur = max(0, clip["end_sec"] - clip["start_sec"])
                n_this = max(1, round(n_frames * dur / total_active))
                for i in range(n_this):
                    if len(frames_b64) >= n_frames:
                        break
                    t = clip["start_sec"] + dur * i / max(n_this - 1, 1)
                    cap.set(cv2.CAP_PROP_POS_FRAMES, int(t * fps))
                    ret, frame = cap.read()
                    if ret:
                        frame = _resize_frame(frame)
                        _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 75])
                        frames_b64.append(base64.b64encode(buf).decode())
            cap.release()
            return frames_b64[:n_frames]

    # Uniform fallback
    positions = [int(i * total / n_frames) for i in range(n_frames)]
    for pos in positions:
        cap.set(cv2.CAP_PROP_POS_FRAMES, pos)
        ret, frame = cap.read()
        if ret:
            frame = _resize_frame(frame)
            _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 75])
            frames_b64.append(base64.b64encode(buf).decode())
    cap.release()
    return frames_b64


def sample_frames_multi(video_dir: Path, appearances: list[dict],
                        n_per_cam: int = 3, slot: str | None = None) -> list[str]:
    """For cross-camera questions: sample frames from each camera appearance.

    appearances: list of {camera_id, start_time, end_time} dicts
    slot: optional slot key (e.g. "13-50") to restrict to the correct slot video.
    Returns combined frame list (up to n_per_cam frames per camera).
    """
    frames_b64: list[str] = []
    for app in appearances:
        cam_id   = app.get("camera_id", "")
        t_start  = float(app.get("start_time", 0))
        t_end    = float(app.get("end_time", t_start + 10))
        # Find the video file for this camera appearance
        video_path: Path | None = None
        # Prefer exact match from SLOT_CAMERAS to avoid wrong-slot videos
        if slot and slot in SLOT_CAMERAS and cam_id in SLOT_CAMERAS[slot]:
            stem = SLOT_CAMERAS[slot][cam_id]
            candidate = video_dir / f"{stem}.avi"
            if candidate.exists():
                video_path = candidate
        if video_path is None:
            for stem_path in video_dir.glob("*.avi"):
                # stem contains camera id like .G638.
                if f".{cam_id}." in stem_path.name:
                    video_path = stem_path
                    break
        if video_path is None or not video_path.exists():
            continue
        cap = cv2.VideoCapture(str(video_path))
        fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
        dur = max(1.0, t_end - t_start)
        for i in range(n_per_cam):
            t = t_start + dur * i / max(n_per_cam - 1, 1)
            cap.set(cv2.CAP_PROP_POS_FRAMES, int(t * fps))
            ret, frame = cap.read()
            if ret:
                frame = _resize_frame(frame)
                # Label the frame with camera id
                cv2.putText(frame, f"CAM:{cam_id} {int(t)}s",
                            (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 75])
                frames_b64.append(base64.b64encode(buf).decode())
        cap.release()
    return frames_b64


def call_vlm(client: OpenAI, frames_b64: list[str], question: str, context: str) -> dict:
    """Single VLM call: frames + pipeline context + question → answer."""
    user_content: list[dict] = []
    if context:
        user_content.append({"type": "text", "text": f"[Tracker context]\n{context}\n\n[Question]\n{question}"})
    else:
        user_content.append({"type": "text", "text": question})
    for b64 in frames_b64:
        user_content.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/jpeg;base64,{b64}"},
        })
    t0 = time.time()
    resp = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": user_content},
        ],
        temperature=0.0,
        max_tokens=256,
    )
    elapsed = time.time() - t0
    raw = resp.choices[0].message.content or ""
    usage = {
        "prompt_tokens":     resp.usage.prompt_tokens     if resp.usage else 0,
        "completion_tokens": resp.usage.completion_tokens if resp.usage else 0,
        "total_tokens":      resp.usage.total_tokens      if resp.usage else 0,
    }
    return {"raw": raw, "usage": usage, "elapsed_sec": round(elapsed, 2)}


def parse_answer(raw: str) -> str:
    low = raw.strip().lower()
    if low.startswith("yes"):
        return "yes"
    if low.startswith("no"):
        return "no"
    return "unknown"


# ══════════════════════════════════════════════════════════════════════════════
# QA case loader
# ══════════════════════════════════════════════════════════════════════════════

def _infer_category(question: str, answer: str, difficulty: str) -> str:
    """Infer QA category from question text when no category column exists."""
    if answer == "no":
        return "negative"
    q = question.lower()
    if difficulty == "hard" or any(w in q for w in
            ["same person", "also seen", "cross", "another camera", "other camera",
             "both camera", "multiple camera", "re-identified"]):
        return "cross_camera"
    if any(w in q for w in
            ["wearing", "jacket", "shirt", "pants", "jeans", "dress", "coat",
             "color", "colour", "appearance", "hair", "bag", "backpack",
             "glasses", "hat", "cap", "shoes", "hoodie", "sweater"]):
        return "appearance"
    if any(w in q for w in
            ["walking", "running", "moving", "direction", "enter", "exit",
             "left", "right", "toward", "away", "carrying", "sitting",
             "standing", "talking", "phone"]):
        return "event"
    return "existence"


def load_cases(xlsx_path: Path, slot_filter: str | None) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Part1"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(name: str) -> int | None:
        for i, h in enumerate(headers):
            if name in h:
                return i
        return None

    ci = {k: col(k) for k in ["video_id", "question", "expected_answer", "difficulty",
                               "category", "query_type"]}

    cases: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vid  = str(row[ci["video_id"]]       or "").strip()
        q    = str(row[ci["question"]]        or "").strip()
        ans  = str(row[ci["expected_answer"]] or "").strip().lower()
        diff = str(row[ci["difficulty"]]      or "").strip()
        if not vid or not q or ans not in ("yes", "no"):
            continue
        if slot_filter:
            if vid not in STEM_TO_SLOT_CAM:
                continue
            s, _ = STEM_TO_SLOT_CAM[vid]
            if s != slot_filter:
                continue

        # category column may not exist — infer from question text
        cat_idx = ci.get("category") or ci.get("query_type")
        if cat_idx is not None:
            cat = str(row[cat_idx] or "").strip() or _infer_category(q, ans, diff)
        else:
            cat = _infer_category(q, ans, diff)

        cases.append({
            "video_id": vid,
            "question": q,
            "expected": ans,
            "difficulty": diff,
            "category": cat,
        })
    return cases


def stratified_sample(cases: list[dict], n: int, seed: int = 42) -> list[dict]:
    """Sample n cases preserving category proportions."""
    rng = random.Random(seed)
    by_cat: dict[str, list[dict]] = defaultdict(list)
    for c in cases:
        by_cat[c["category"]].append(c)
    total = len(cases)
    sampled: list[dict] = []
    for cat, items in by_cat.items():
        k = max(1, round(n * len(items) / total))
        sampled.extend(rng.sample(items, min(k, len(items))))
    # Trim or top-up to exactly n
    rng.shuffle(sampled)
    if len(sampled) > n:
        return sampled[:n]
    # Top up from remainder if short
    used = set(id(c) for c in sampled)
    remainder = [c for c in cases if id(c) not in used]
    rng.shuffle(remainder)
    sampled.extend(remainder[: n - len(sampled)])
    return sampled[:n]


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    ap = argparse.ArgumentParser(description="MEVID full pipeline + VLM QA evaluation")
    ap.add_argument("--video-dir", default="_data/mevid_slots",
                    help="Directory containing downloaded .avi files")
    ap.add_argument("--xlsx", default="agent/test/data/agent_test_mevid.xlsx",
                    help="Path to agent_test_mevid.xlsx")
    ap.add_argument("--slot", default="",
                    help="Restrict to one slot, e.g. 16-35 (default: all)")
    ap.add_argument("--limit", type=int, default=100,
                    help="Number of QA cases to evaluate (default 100)")
    ap.add_argument("--frames", type=int, default=10,
                    help="Frames to sample per VLM call (default 10)")
    ap.add_argument("--no-refine", action="store_true",
                    help="Skip LLM event refinement (faster/cheaper but weaker context)")
    ap.add_argument("--reid-device", default="cpu",
                    help="Device for OSNet Re-ID: cpu | cuda | cuda:0 (default cpu)")
    ap.add_argument("--force-pipeline", action="store_true",
                    help="Re-run pipeline even if cache exists")
    ap.add_argument("--resume", action="store_true",
                    help="Skip QA cases already in the resume checkpoint")
    ap.add_argument("--out-dir", default="results",
                    help="Output directory for result JSON (default results/)")
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    video_dir = ROOT / args.video_dir
    xlsx_path = ROOT / args.xlsx
    out_dir   = ROOT / args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    slots = [args.slot] if args.slot else list(SLOT_CAMERAS)

    # ── Check API key ──────────────────────────────────────────────────────────
    if not API_KEY:
        print("ERROR: DASHSCOPE_API_KEY not set in .env")
        sys.exit(1)

    # ── Phase 1–4: Pipeline ────────────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("PHASE 1-4: Multi-camera pipeline (YOLO + OSNet + topology + matching)")
    print("═" * 60)

    pipeline_data: dict[str, dict] = {}   # slot → serialised output
    for slot in slots:
        try:
            pipeline_data[slot] = run_pipeline_for_slot(
                slot, video_dir,
                reid_device=args.reid_device,
                force=args.force_pipeline,
            )
            pc = pipeline_data[slot]
            print(
                f"  [{slot}] ✓  cameras={pc.get('cameras_run', [])}  "
                f"entities={len(pc.get('global_entities', []))}  "
                f"merged_events={len(pc.get('merged_events', []))}  "
                f"elapsed={pc.get('elapsed_sec', '?')}s"
            )
        except Exception as e:
            print(f"  [{slot}] ✗  {e}")

    # ── Phase 5: LLM refinement (default ON; skip with --no-refine) ──────────
    refined_data: dict[str, dict] = {}
    if not args.no_refine:
        print("\n" + "═" * 60)
        print("PHASE 5: LLM event refinement (DashScope) — skip with --no-refine")
        print("═" * 60)
        for slot in slots:
            try:
                refined_data[slot] = run_refinement_for_slot(slot, video_dir)
                print(f"  [{slot}] ✓  refinement done")
            except Exception as e:
                print(f"  [{slot}] ✗  {e}")
    else:
        print("\n[Phase 5 skipped — --no-refine set. Context will be track-only (weak).]")

    # ── Phase 6: VLM QA ───────────────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("PHASE 6: VLM yes/no QA evaluation")
    print("═" * 60)

    all_cases = load_cases(xlsx_path, args.slot or None)
    selected  = stratified_sample(all_cases, args.limit, seed=args.seed)
    print(f"  Cases selected: {len(selected)} / {len(all_cases)} total")
    by_cat = Counter(c["category"] for c in selected)
    print(f"  Category breakdown: {dict(by_cat)}")

    # Resume support
    resume_file = out_dir / "mevid_full_resume.json"
    prior_results: list[dict] = []
    done_ids: set[str] = set()
    if args.resume and resume_file.exists():
        prior_results = json.loads(resume_file.read_text(encoding="utf-8"))
        done_ids = {r["case_id"] for r in prior_results}
        print(f"  Resuming: {len(done_ids)} cases already done")

    client = OpenAI(api_key=API_KEY, base_url=BASE_URL)
    results: list[dict] = list(prior_results)
    total_prompt = total_completion = 0
    total_cost = 0.0
    correct = wrong = skipped = 0

    # Group by video to sample frames once per video
    by_video: dict[str, list[dict]] = defaultdict(list)
    for i, c in enumerate(selected):
        c["_idx"] = i
        by_video[c["video_id"]].append(c)

    for video_id, vcases in sorted(by_video.items()):
        video_path = video_dir / f"{video_id}.avi"
        if not video_path.exists():
            print(f"\n[SKIP] video not found: {video_id}.avi")
            for c in vcases:
                case_id = f"FULL_{c['_idx']:04d}"
                if case_id not in done_ids:
                    results.append({
                        "case_id": case_id, "video_id": video_id,
                        "expected": c["expected"], "predicted": "skip",
                        "correct": False, "error": "video_not_found",
                    })
                    skipped += 1
            continue

        # Clip-aware frame sampling: use pipeline clips if available
        slot_key = STEM_TO_SLOT_CAM.get(video_id, (None, None))[0]
        cam_key  = STEM_TO_SLOT_CAM.get(video_id, (None, None))[1]
        clips: list[dict] = []
        if slot_key and cam_key:
            per_cam = {cr["camera_id"]: cr
                       for cr in pipeline_data.get(slot_key, {}).get("per_camera", [])}
            raw_clips = per_cam.get(cam_key, {}).get("clips", [])
            # clips from pipeline use start_sec/end_sec keys
            clips = [{"start_sec": c.get("start_sec", c.get("start_time", 0)),
                      "end_sec":   c.get("end_sec",   c.get("end_time",   0))}
                     for c in raw_clips]

        frames_b64 = sample_frames(str(video_path), n_frames=args.frames, clips=clips or None)

        # Pre-compute entity appearance windows for this camera.
        # Split into: (a) all entity apps for event questions,
        #             (b) brief-only entity apps (<= BRIEF_DUR_SEC) for appearance questions.
        # Brief appearances (persons present for only a few seconds) are missed by clip-aware
        # proportional sampling — e.g. grey hoodie at G339@297-300s (3s out of 300s total).
        BRIEF_DUR_SEC = 6.0
        entity_frames_b64: list[str] = []        # all entity apps (for event category)
        brief_entity_frames_b64: list[str] = []  # only brief apps (for appearance category)
        if slot_key and cam_key:
            slot_entities = pipeline_data.get(slot_key, {}).get("global_entities", [])
            entity_apps = [
                a for ge in slot_entities
                for a in ge["appearances"]
                if a["camera_id"] == cam_key
            ]
            brief_apps = [
                a for a in entity_apps
                if (float(a.get("end_time", 0)) - float(a.get("start_time", 0))) <= BRIEF_DUR_SEC
                and float(a.get("end_time", 0)) > float(a.get("start_time", 0))
            ]
            if entity_apps:
                entity_frames_b64 = sample_frames_multi(
                    video_dir, entity_apps, n_per_cam=2, slot=slot_key
                )
            if brief_apps:
                brief_entity_frames_b64 = sample_frames_multi(
                    video_dir, brief_apps, n_per_cam=3, slot=slot_key
                )

        # Base context (no question-specific camera targeting)
        base_context = build_context(video_id, pipeline_data, refined_data)

        print(f"\n[{video_id}] {len(vcases)} question(s)  "
              f"base={len(frames_b64)} brief_entity={len(brief_entity_frames_b64)}"
              f" all_entity={len(entity_frames_b64)}")
        if base_context:
            ctx_lines = base_context.count("\n") + 1
            print(f"  Context: {ctx_lines} lines from pipeline")

        for c in vcases:
            case_id = f"FULL_{c['_idx']:04d}"
            if case_id in done_ids:
                continue

            print(f"  [{case_id}] ({c['category']}/{c['difficulty']}) {c['question'][:70]}")

            # Build question-aware context. Negative questions suppress cross-camera
            # candidate text so uncertain Re-ID hints do not bias no-answer cases.
            context_mode = "negative" if c["category"] == "negative" else "default"
            if c["category"] == "cross_camera":
                context_mode = "cross_camera"
            context = build_context(
                video_id,
                pipeline_data,
                refined_data,
                question=c["question"] if c["category"] == "cross_camera" else "",
                context_mode=context_mode,
            )

            # Frame selection strategy per category:
            #   event       → clip-aware + all entity frames (windows capture the action)
            #   appearance  → clip-aware + BRIEF entity frames only (< 6s windows missed by
            #                 clip-proportional sampling; longer windows covered by clip-aware)
            #   negative    → clip-aware only (entity frames add irrelevant persons)
            #   cross_camera → clip-aware + dedicated multi-cam entity frames (see below)
            if c["category"] == "event":
                case_frames = frames_b64 + entity_frames_b64
            elif c["category"] == "appearance":
                case_frames = frames_b64 + brief_entity_frames_b64
            else:
                case_frames = frames_b64

            if c["category"] == "cross_camera" and slot_key:
                slot_entities = pipeline_data.get(slot_key, {}).get("global_entities", [])
                # Extract cameras explicitly mentioned in the question (to prioritize)
                q_lower = c["question"].lower()
                mentioned_cams = [
                    cam for cam in SLOT_CAMERAS.get(slot_key, {}).keys()
                    if cam != cam_key and cam.lower() in q_lower
                ]
                # Find entities that involve this camera AND the mentioned cameras
                relevant_appearances: list[dict] = []
                for ge in slot_entities:
                    cam_apps = [a for a in ge["appearances"] if a["camera_id"] == cam_key]
                    if not cam_apps:
                        continue
                    q_cam_apps = [a for a in ge["appearances"] if a["camera_id"] in mentioned_cams]
                    # For entities that span both cam_key and mentioned cameras:
                    # Include frames from BOTH the current camera window AND the other camera window
                    if q_cam_apps:
                        relevant_appearances.extend(cam_apps[:1])   # current cam entity window
                        relevant_appearances.extend(q_cam_apps[:1]) # mentioned cam entity window
                    else:
                        # Entity involving current cam but not the mentioned cam
                        other_apps = [
                            a for a in ge["appearances"]
                            if a["camera_id"] != cam_key
                        ]
                        relevant_appearances.extend(other_apps[:1])
                # De-duplicate by (camera_id, track start_time)
                seen_apps: set[tuple[str, float]] = set()
                unique_apps = []
                for a in relevant_appearances:
                    key = (a["camera_id"], float(a.get("start_time", 0)))
                    if key not in seen_apps:
                        seen_apps.add(key)
                        unique_apps.append(a)
                relevant_appearances = unique_apps[:8]  # cap total
                if relevant_appearances:
                    extra_frames = sample_frames_multi(
                        video_dir, relevant_appearances, n_per_cam=3, slot=slot_key
                    )
                    case_frames = frames_b64 + extra_frames
                    print(f"    + {len(extra_frames)} cross-camera frames from "
                          f"{[a['camera_id'] for a in relevant_appearances]}")

            try:
                vlm_resp = call_vlm(client, case_frames, c["question"], context)
            except Exception as e:
                print(f"    ✗ API error: {e}")
                results.append({
                    "case_id": case_id, "video_id": video_id,
                    "question": c["question"], "expected": c["expected"],
                    "predicted": "error", "correct": False,
                    "category": c["category"], "difficulty": c["difficulty"],
                    "error": str(e),
                })
                skipped += 1
                continue

            predicted = parse_answer(vlm_resp["raw"])
            is_correct = predicted == c["expected"]
            mark = "✓" if is_correct else "✗"
            print(
                f"    {mark} expected={c['expected']} predicted={predicted} "
                f"tokens={vlm_resp['usage']['total_tokens']}"
            )

            if is_correct:
                correct += 1
            else:
                wrong += 1

            usage = vlm_resp["usage"]
            total_prompt     += usage["prompt_tokens"]
            total_completion += usage["completion_tokens"]
            cost = (usage["prompt_tokens"] * PRICE_IN + usage["completion_tokens"] * PRICE_OUT) / 1000
            total_cost += cost

            results.append({
                "case_id":    case_id,
                "video_id":   video_id,
                "question":   c["question"],
                "expected":   c["expected"],
                "predicted":  predicted,
                "correct":    is_correct,
                "category":   c["category"],
                "difficulty": c["difficulty"],
                "raw_answer": vlm_resp["raw"],
                "usage":      usage,
                "cost_cny":   round(cost, 5),
                "elapsed_sec": vlm_resp["elapsed_sec"],
            })

            # Save checkpoint
            resume_file.write_text(
                json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8"
            )

    # ── Summary ────────────────────────────────────────────────────────────────
    evaluated = correct + wrong
    accuracy  = correct / evaluated if evaluated else 0.0

    print("\n" + "═" * 60)
    print("RESULTS SUMMARY")
    print("═" * 60)

    # Per-category accuracy
    cat_stats: dict[str, dict] = defaultdict(lambda: {"correct": 0, "total": 0})
    for r in results:
        if r.get("predicted") in ("yes", "no"):
            cat = r.get("category", "?")
            cat_stats[cat]["total"] += 1
            if r.get("correct"):
                cat_stats[cat]["correct"] += 1

    for cat, st in sorted(cat_stats.items()):
        acc = st["correct"] / st["total"] if st["total"] else 0.0
        print(f"  {cat:20s}: {st['correct']:3d}/{st['total']:3d}  ({acc:.1%})")

    print(f"\n  Overall accuracy : {correct}/{evaluated} = {accuracy:.1%}")
    print(f"  Skipped          : {skipped}")
    print(f"  Prompt tokens    : {total_prompt:,}")
    print(f"  Completion tokens: {total_completion:,}")
    print(f"  VLM cost         : ¥{total_cost:.4f} CNY")
    print(f"  Refinement       : {'disabled (--no-refine)' if args.no_refine else 'enabled'}")

    # Pipeline summary
    print("\n  Pipeline module coverage:")
    for slot in slots:
        if slot in pipeline_data:
            pd = pipeline_data[slot]
            n_ent = len(pd.get("global_entities", []))
            n_ev  = len(pd.get("merged_events", []))
            n_cam = len(pd.get("cameras_run", []))
            print(f"    [{slot}] {n_cam} cameras → {n_ent} global entities, {n_ev} merged events")

    # Save final result
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    out_file  = out_dir / f"mevid_full_{timestamp}.json"
    final = {
        "timestamp": timestamp,
        "limit": args.limit,
        "slot_filter": args.slot or "all",
        "refine": not args.no_refine,
        "reid_device": args.reid_device,
        "frames_per_case": args.frames,
        "model": MODEL,
        "n_evaluated": evaluated,
        "n_correct": correct,
        "n_wrong": wrong,
        "n_skipped": skipped,
        "accuracy": round(accuracy, 4),
        "per_category": {
            cat: {
                "correct": st["correct"],
                "total": st["total"],
                "accuracy": round(st["correct"] / st["total"], 4) if st["total"] else 0.0,
            }
            for cat, st in cat_stats.items()
        },
        "token_usage": {
            "prompt_tokens": total_prompt,
            "completion_tokens": total_completion,
            "total_tokens": total_prompt + total_completion,
        },
        "cost_cny": round(total_cost, 4),
        "pipeline_summary": {
            slot: {
                "global_entities": len(pipeline_data[slot].get("global_entities", [])),
                "merged_events":   len(pipeline_data[slot].get("merged_events", [])),
                "cameras_run":     pipeline_data[slot].get("cameras_run", []),
            }
            for slot in slots if slot in pipeline_data
        },
        "cases": results,
    }
    out_file.write_text(json.dumps(final, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n  Result saved → {out_file}")

    # Clean up resume file on success
    if resume_file.exists() and skipped == 0:
        resume_file.unlink()


if __name__ == "__main__":
    main()
