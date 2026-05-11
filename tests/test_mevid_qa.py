"""
tests/test_mevid_qa.py
-----------------------
Runs the MEVID multi-camera QA evaluation.

Pipeline per test case:
  1. Load the target video file
  2. Run YOLO+BoT-SORT to detect motion clips  (results cached per video)
  3. Sample frames from the motion clips
  4. Call qwen-vl-max-latest with the question
  5. Parse yes/no answer + optional time range
  6. Compare against expected_answer
  7. Save per-case results and aggregate metrics to results/mevid_qa_*.json

Usage:
    # full run
    python tests/test_mevid_qa.py

    # quick smoke test (5 cases)
    python tests/test_mevid_qa.py --limit 5

    # specific category only
    python tests/test_mevid_qa.py --category cross_camera

    # choose input xlsx / video dir
    python tests/test_mevid_qa.py \\
        --xlsx agent/test/data/agent_test_mevid.xlsx \\
        --video-dir _data/mevid_slots

    # resume from previous partial run
    python tests/test_mevid_qa.py --resume
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path

import cv2

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

# Load .env
try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_ROOT / ".env")
except ImportError:
    pass

from openai import OpenAI

# ── Defaults ──────────────────────────────────────────────────────────────────
DEFAULT_XLSX   = PROJECT_ROOT / "agent" / "test" / "data" / "agent_test_mevid.xlsx"
DEFAULT_VIDDIR = PROJECT_ROOT / "_data" / "mevid_slots"
DEFAULT_OUT    = PROJECT_ROOT / "results"
YOLO_CACHE_DIR = PROJECT_ROOT / "_cache" / "yolo_mevid"

MODEL   = os.environ.get("DASHSCOPE_CHAT_MODEL", "qwen-vl-max-latest")
API_KEY = os.environ.get("DASHSCOPE_API_KEY", "")
BASE_URL = os.environ.get("DASHSCOPE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1")

# ── Import project utilities ───────────────────────────────────────────────────
try:
    from ultralytics import YOLO
    YOLO_AVAILABLE = True
except ImportError:
    YOLO_AVAILABLE = False
    print("[WARN] ultralytics not installed; will use uniform frame sampling instead of YOLO")

# ── Load test cases from xlsx ─────────────────────────────────────────────────
def load_cases(xlsx_path: Path, category: str | None = None) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Part1" not in wb.sheetnames:
        raise ValueError(f"No 'Part1' sheet in {xlsx_path}")
    ws = wb["Part1"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 1 = headers
    headers = [str(v or "").strip().lower() for v in rows[0]]

    def col(row, name):
        try:
            i = headers.index(name)
            return str(row[i] or "").strip() if i < len(row) else ""
        except ValueError:
            return ""

    cases = []
    current_vid = ""
    for row in rows[1:]:
        if not any(v for v in row):
            continue
        vid = col(row, "video_id") or current_vid
        if col(row, "video_id"):
            current_vid = vid
        q = col(row, "question")
        if not q:
            continue
        cases.append({
            "video_id":        vid,
            "question":        q,
            "recall_challenge": col(row, "recall_challenge"),
            "expected_answer": col(row, "expected_answer").lower(),
            "expected_time":   col(row, "expected_time"),
            "difficulty":      col(row, "difficulty"),
            "_category":       col(row, "recall_challenge"),  # approximation
        })

    # Filter by category if requested (we use the recall_challenge as a proxy;
    # the category label is not stored in xlsx — use video_id + question heuristics)
    if category:
        cat_map = {
            "existence":    lambda c: c["question"].lower().startswith("is there a person visible"),
            "appearance":   lambda c: "wearing" in c["question"].lower() or "carrying" in c["question"].lower(),
            "event":        lambda c: "exit from" in c["question"].lower(),
            "cross_camera": lambda c: "appear again in camera" in c["question"].lower(),
            "negative":     lambda c: c["expected_answer"] == "no",
        }
        fn = cat_map.get(category)
        if fn:
            cases = [c for c in cases if fn(c)]
        else:
            print(f"[WARN] unknown category filter '{category}', ignored")

    return cases


# ── YOLO motion clip detection ────────────────────────────────────────────────
YOLO_MODEL_PATH = PROJECT_ROOT / "_model" / "yolo11m.pt"

def get_motion_clips(video_path: str, cache_dir: Path) -> list[dict]:
    """Run YOLO+BoT-SORT and return motion clip segments.
    Results are cached to avoid re-processing the same video.
    """
    stem = Path(video_path).stem
    cache_file = cache_dir / f"{stem}_clips.json"
    if cache_file.exists():
        return json.loads(cache_file.read_text(encoding="utf-8"))

    if not YOLO_AVAILABLE:
        # Fallback: treat whole video as one clip
        cap = cv2.VideoCapture(video_path)
        fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
        n   = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        cap.release()
        dur = n / fps
        clips = [{"start_sec": 0.0, "end_sec": dur, "n_tracks": 0}]
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_file.write_text(json.dumps(clips), encoding="utf-8")
        return clips

    model = YOLO(str(YOLO_MODEL_PATH))
    PERSON_CLASS = 0
    conf, iou = 0.25, 0.25

    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
    cap.release()

    results = model.track(
        video_path,
        classes=[PERSON_CLASS],
        conf=conf,
        iou=iou,
        tracker="botsort.yaml",
        stream=True,
        verbose=False,
    )

    active_frames: list[int] = []
    for frame_idx, r in enumerate(results):
        if r.boxes is not None and len(r.boxes) > 0:
            active_frames.append(frame_idx)

    if not active_frames:
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_file.write_text(json.dumps([]), encoding="utf-8")
        return []

    # Merge nearby active frames into clips (gap ≤ 2 s)
    gap_frames = int(fps * 2)
    clips: list[dict] = []
    seg_start = active_frames[0]
    seg_end   = active_frames[0]
    for f in active_frames[1:]:
        if f - seg_end <= gap_frames:
            seg_end = f
        else:
            clips.append({"start_sec": seg_start / fps, "end_sec": seg_end / fps})
            seg_start = seg_end = f
    clips.append({"start_sec": seg_start / fps, "end_sec": seg_end / fps})

    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file.write_text(json.dumps(clips), encoding="utf-8")
    return clips


# ── Frame sampling ─────────────────────────────────────────────────────────────
def sample_frames(video_path: str, clips: list[dict], n_frames: int = 10) -> list[str]:
    """Sample n_frames from motion clips; return list of base64 JPEG strings."""
    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    if not clips:
        # Uniform fallback across full video
        positions = [int(i * total_frames / n_frames) for i in range(n_frames)]
    else:
        # Spread evenly over all motion clips
        total_sec = sum(c["end_sec"] - c["start_sec"] for c in clips)
        frames_b64: list[str] = []
        for clip in clips:
            dur = clip["end_sec"] - clip["start_sec"]
            n_this = max(1, round(n_frames * dur / max(total_sec, 1e-6)))
            for i in range(n_this):
                t = clip["start_sec"] + dur * i / max(n_this - 1, 1)
                positions = [int(t * fps)]
                if len(frames_b64) >= n_frames:
                    break
                cap.set(cv2.CAP_PROP_POS_FRAMES, positions[0])
                ret, frame = cap.read()
                if ret:
                    _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
                    frames_b64.append(base64.b64encode(buf).decode())
            if len(frames_b64) >= n_frames:
                break
        cap.release()
        return frames_b64[:n_frames]

    frames_b64 = []
    for pos in positions[:n_frames]:
        cap.set(cv2.CAP_PROP_POS_FRAMES, pos)
        ret, frame = cap.read()
        if ret:
            _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
            frames_b64.append(base64.b64encode(buf).decode())
    cap.release()
    return frames_b64


# ── VLM call ──────────────────────────────────────────────────────────────────
SYSTEM_PROMPT = (
    "You are a surveillance video analyst. "
    "You will be shown frames from a security camera clip (each ~5 minutes long). "
    "Answer the user's question about what appears in the video. "
    "Your answer MUST start with 'yes' or 'no', followed by a brief explanation. "
    "If yes, include the approximate time range (mm:ss format) when the subject is visible. "
    "Be concise. Do not invent details not visible in the frames."
)


def call_vlm(client: OpenAI, frames_b64: list[str], question: str) -> dict:
    content = [{"type": "text", "text": question}]
    for b64 in frames_b64:
        content.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/jpeg;base64,{b64}"},
        })
    t0 = time.time()
    resp = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": content},
        ],
        temperature=0.0,
        max_tokens=256,
    )
    elapsed = time.time() - t0
    raw = resp.choices[0].message.content or ""
    usage = {
        "prompt_tokens":     resp.usage.prompt_tokens     if resp.usage else 0,
        "completion_tokens": resp.usage.completion_tokens if resp.usage else 0,
        "total_tokens":      resp.usage.total_tokens      if resp.usage else 0,
    }
    return {"raw": raw, "usage": usage, "elapsed_sec": round(elapsed, 2)}


def parse_answer(raw: str) -> tuple[str, str | None]:
    """Returns (answer_label, time_range_str|None)."""
    low = raw.strip().lower()
    label = "unknown"
    if low.startswith("yes"):
        label = "yes"
    elif low.startswith("no"):
        label = "no"

    # Extract time range like "0:28-1:15" or "around 2:30"
    import re
    m = re.search(r'(\d+:\d+)\s*[-–]\s*(\d+:\d+)', raw)
    if m:
        return label, f"{m.group(1)}-{m.group(2)}"
    m2 = re.search(r'(\d+:\d+)', raw)
    if m2:
        return label, m2.group(1)
    return label, None


# ── Main evaluation loop ──────────────────────────────────────────────────────
def run_evaluation(args: argparse.Namespace) -> None:
    xlsx_path = Path(args.xlsx)
    video_dir = Path(args.video_dir)
    out_dir   = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    out_file  = out_dir / f"mevid_qa_{timestamp}.json"

    print(f"Loading test cases from: {xlsx_path}")
    cases = load_cases(xlsx_path, args.category or None)
    if args.limit:
        cases = cases[: args.limit]
    print(f"  {len(cases)} cases selected")

    # Resume support: skip already-done cases
    done_ids: set[str] = set()
    resume_file = out_dir / "mevid_qa_resume.json"
    prior_results: list[dict] = []
    if args.resume and resume_file.exists():
        prior_results = json.loads(resume_file.read_text(encoding="utf-8"))
        done_ids = {r["case_id"] for r in prior_results}
        print(f"  Resuming: {len(done_ids)} cases already done, skipping them")

    if not API_KEY:
        print("ERROR: DASHSCOPE_API_KEY not set in .env")
        sys.exit(1)

    client = OpenAI(api_key=API_KEY, base_url=BASE_URL)
    PRICE_IN  = float(os.environ.get("QWEN_PRICE_IN",  "0.012"))   # ¥ per 1k tokens
    PRICE_OUT = float(os.environ.get("QWEN_PRICE_OUT", "0.036"))

    # Pre-process: group cases by video to avoid re-running YOLO per case
    by_video: dict[str, list[dict]] = defaultdict(list)
    for i, c in enumerate(cases):
        c["_idx"] = i
        by_video[c["video_id"]].append(c)

    results: list[dict] = list(prior_results)
    total_tokens = 0
    total_cost   = 0.0
    correct = wrong = skipped = 0

    for video_id, vcases in sorted(by_video.items()):
        # Find video file
        video_path = video_dir / f"{video_id}.avi"
        if not video_path.exists():
            print(f"\n[SKIP] video not found: {video_path}")
            for c in vcases:
                case_id = f"PART1_{c['_idx']:04d}"
                if case_id not in done_ids:
                    results.append({
                        "case_id": case_id,
                        "video_id": video_id,
                        "question": c["question"],
                        "expected": c["expected_answer"],
                        "predicted": "skip",
                        "correct": False,
                        "error": "video_not_found",
                    })
                    skipped += 1
            continue

        print(f"\n{'='*60}")
        print(f"Video: {video_id}  ({len(vcases)} questions)")
        print(f"File:  {video_path}")

        # Run YOLO once per video
        clips = get_motion_clips(str(video_path), YOLO_CACHE_DIR)
        print(f"  Motion clips: {len(clips)}")

        # Sample frames once per video (all questions share the same frames)
        frames_b64 = sample_frames(str(video_path), clips, n_frames=args.frames)
        print(f"  Frames sampled: {len(frames_b64)}")

        for c in vcases:
            case_id = f"PART1_{c['_idx']:04d}"
            if case_id in done_ids:
                continue

            print(f"\n  [{case_id}] {c['question'][:80]}")
            print(f"           expected={c['expected_answer']}  difficulty={c['difficulty']}")

            try:
                vlm = call_vlm(client, frames_b64, c["question"])
                predicted_label, pred_time = parse_answer(vlm["raw"])
                usage = vlm["usage"]
                cost = (usage["prompt_tokens"] * PRICE_IN
                        + usage["completion_tokens"] * PRICE_OUT) / 1000
                total_tokens += usage["total_tokens"]
                total_cost   += cost

                is_correct = (predicted_label == c["expected_answer"])
                correct += int(is_correct)
                wrong   += int(not is_correct)

                print(f"           predicted={predicted_label}  "
                      f"{'✓' if is_correct else '✗'}  "
                      f"tokens={usage['total_tokens']}  ¥{cost:.3f}  "
                      f"({vlm['elapsed_sec']}s)")

                results.append({
                    "case_id":          case_id,
                    "video_id":         video_id,
                    "question":         c["question"],
                    "expected":         c["expected_answer"],
                    "predicted":        predicted_label,
                    "predicted_time":   pred_time,
                    "expected_time":    c["expected_time"],
                    "correct":          is_correct,
                    "difficulty":       c["difficulty"],
                    "recall_challenge": c["recall_challenge"],
                    "vlm_raw":          vlm["raw"],
                    "usage":            usage,
                    "cost_cny":         round(cost, 4),
                    "elapsed_sec":      vlm["elapsed_sec"],
                    "error":            None,
                })

            except Exception as e:
                print(f"           ERROR: {e}")
                results.append({
                    "case_id":   case_id,
                    "video_id":  video_id,
                    "question":  c["question"],
                    "expected":  c["expected_answer"],
                    "predicted": "error",
                    "correct":   False,
                    "error":     str(e),
                })
                skipped += 1

            # Save checkpoint after each case
            resume_file.write_text(
                json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8"
            )

    # ── Aggregate metrics ──────────────────────────────────────────────────────
    evaluated = correct + wrong
    accuracy  = round(correct / max(evaluated, 1), 4)

    # Per-category breakdown
    cat_stats: dict[str, dict] = {}
    for r in results:
        if r.get("error") or r.get("predicted") in ("skip", "error"):
            continue
        q = r["question"].lower()
        if "appear again in camera" in q:
            cat = "cross_camera"
        elif "exit from" in q:
            cat = "event"
        elif "wearing" in q or "carrying" in q:
            cat = "appearance"
        elif r["expected"] == "no":
            cat = "negative"
        else:
            cat = "existence"
        if cat not in cat_stats:
            cat_stats[cat] = {"correct": 0, "total": 0}
        cat_stats[cat]["total"] += 1
        cat_stats[cat]["correct"] += int(r.get("correct", False))
    for cat in cat_stats:
        n = cat_stats[cat]["total"]
        c_ = cat_stats[cat]["correct"]
        cat_stats[cat]["accuracy"] = round(c_ / max(n, 1), 4)

    # Yes/No breakdown
    yes_total  = sum(1 for r in results if r.get("expected") == "yes" and r.get("predicted") != "skip")
    yes_correct = sum(1 for r in results if r.get("expected") == "yes" and r.get("correct"))
    no_total   = sum(1 for r in results if r.get("expected") == "no"  and r.get("predicted") != "skip")
    no_correct  = sum(1 for r in results if r.get("expected") == "no"  and r.get("correct"))

    summary = {
        "total_cases":    len(cases),
        "evaluated":      evaluated,
        "skipped":        skipped,
        "correct":        correct,
        "wrong":          wrong,
        "accuracy":       accuracy,
        "yes_accuracy":   round(yes_correct / max(yes_total, 1), 4),
        "no_accuracy":    round(no_correct  / max(no_total,  1), 4),
        "yes_total":      yes_total,
        "no_total":       no_total,
        "total_tokens":   total_tokens,
        "total_cost_cny": round(total_cost, 4),
        "model":          MODEL,
        "frames_per_video": args.frames,
        "per_category":   cat_stats,
    }

    output = {"summary": summary, "case_results": results}
    out_file.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n{'='*60}")
    print(f"RESULTS  —  {evaluated} evaluated  |  {skipped} skipped")
    print(f"  Accuracy:     {accuracy:.1%}  ({correct}/{evaluated})")
    print(f"  Yes accuracy: {summary['yes_accuracy']:.1%}  ({yes_correct}/{yes_total})")
    print(f"  No  accuracy: {summary['no_accuracy']:.1%}  ({no_correct}/{no_total})")
    print(f"  Total tokens: {total_tokens:,}   Cost: ¥{total_cost:.2f}")
    print(f"\nPer-category:")
    for cat, s in sorted(cat_stats.items()):
        print(f"  {cat:<14} {s['accuracy']:.1%}  ({s['correct']}/{s['total']})")
    print(f"\nSaved: {out_file}")

    # Clean up resume file on successful completion
    if not args.resume and resume_file.exists():
        resume_file.unlink(missing_ok=True)


# ── CLI ───────────────────────────────────────────────────────────────────────
def get_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="MEVID multi-camera QA evaluation")
    ap.add_argument("--xlsx",      default=str(DEFAULT_XLSX),
                    help="Path to agent_test_mevid.xlsx")
    ap.add_argument("--video-dir", default=str(DEFAULT_VIDDIR),
                    help="Directory containing the downloaded .avi files")
    ap.add_argument("--out-dir",   default=str(DEFAULT_OUT),
                    help="Directory for result JSON files")
    ap.add_argument("--frames",    type=int, default=10,
                    help="Frames to sample per video (default 10)")
    ap.add_argument("--limit",     type=int, default=0,
                    help="Limit number of cases (0 = all)")
    ap.add_argument("--category",  default="",
                    choices=["", "existence", "appearance", "event", "cross_camera", "negative"],
                    help="Only run one category")
    ap.add_argument("--resume",    action="store_true",
                    help="Skip cases already in results/mevid_qa_resume.json")
    return ap.parse_args()


if __name__ == "__main__":
    run_evaluation(get_args())
